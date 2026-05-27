"""
从 PostgreSQL 导出经验库两张表：experiences + design_lessons
输出：每个表各一套 CSV / JSON，以及合并的 Excel（两个 sheet）
"""
import csv, json, os
from datetime import datetime

try:
    import psycopg2
    import openpyxl
except ImportError:
    print("正在安装依赖…")
    os.system("pip3 install psycopg2-binary openpyxl --user -q")
    import psycopg2
    import openpyxl

OUTPUT_DIR = r"D:\my_crew\outputs\export"
os.makedirs(OUTPUT_DIR, exist_ok=True)
SCHEMA = "8775_T1Q_国内"
DB_DSN = "host=localhost port=5433 dbname=blackscreen user=uidq2071 password=BlackScreen@2025"

TABLES = [
    ("experiences", "经验库"),
    ("design_lessons", "设计经验"),
]

CAT_COLORS = {
    "应用crash": "FFCCCC", "QNX": "CCE5FF", "硬件": "FFE5CC",
    "系统": "D9F2D9", "场景": "FFFFCC", "环境": "E5CCFF",
    "应用": "CCFFFF", "需人工判断": "FFD9E6",
}


def cell_val(v):
    """清理单元格值：列表→字符串，时区时间→无时区"""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    return v


def write_csv(name, cols, rows):
    path = os.path.join(OUTPUT_DIR, f"{name}.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for r in rows:
            w.writerow([cell_val(v) for v in r])
    print(f"✅ CSV:   {path}  ({os.path.getsize(path)/1024:.0f} KB)")


def write_json(name, cols, rows):
    path = os.path.join(OUTPUT_DIR, f"{name}.json")
    data = []
    for r in rows:
        d = {}
        for c, v in zip(cols, r):
            v = cell_val(v)
            if isinstance(v, datetime):
                v = v.strftime("%Y-%m-%d %H:%M:%S")
            d[c] = v
        data.append(d)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON:  {path}  ({os.path.getsize(path)/1024:.0f} KB)")


def write_sheet(ws, name, cols, rows, color_col=None):
    ws.append(cols)
    from openpyxl.styles import PatternFill
    for r_idx, r in enumerate(rows, 2):
        for c_idx, v in enumerate(r, 1):
            v = cell_val(v)
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            if color_col is not None and c_idx == color_col and v in CAT_COLORS:
                c.fill = PatternFill(start_color=CAT_COLORS[v], end_color=CAT_COLORS[v], fill_type="solid")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)


# === 连接数据库 ===
conn = psycopg2.connect(DB_DSN, connect_timeout=5)
wb = openpyxl.Workbook()
# 删默认 sheet
wb.remove(wb.active)

for table_name, sheet_name in TABLES:
    cur = conn.cursor()
    cur.execute(f'SELECT * FROM "{SCHEMA}"."{table_name}" ORDER BY id')
    cols = [desc[0] for desc in cur.description]
    rows = cur.fetchall()
    cur.close()
    print(f"\n「{table_name}」共 {len(rows)} 条，{len(cols)} 个字段")
    print(f"  字段：{', '.join(cols)}")

    # CSV + JSON
    write_csv(table_name, cols, rows)
    write_json(table_name, cols, rows)

    # Excel sheet
    color_col = 2 if table_name == "experiences" else None  # experiences 第2列是分类，上色
    ws = wb.create_sheet(title=sheet_name)
    write_sheet(ws, table_name, cols, rows, color_col=color_col)

conn.close()

xlsx_path = os.path.join(OUTPUT_DIR, "经验库_导出.xlsx")
wb.save(xlsx_path)
print(f"\n✅ Excel: {xlsx_path}  ({os.path.getsize(xlsx_path)/1024:.0f} KB)")
print("\n输出目录：", OUTPUT_DIR)
