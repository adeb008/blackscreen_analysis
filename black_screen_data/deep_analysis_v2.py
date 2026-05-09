#!/usr/bin/env python
"""黑卡闪问题深度根因分析 v2 —— 改进提取 + 对策合理性判断"""

import openpyxl
import re
from collections import Counter, defaultdict

EXCEL_PATH = r"D:\my_crew\black_screen_data\Bug_20260508171433.xlsx"
OUTPUT_PATH = r"D:\my_crew\black_screen_data\deep_analysis_v2.md"


def extract_structured(text: str) -> dict:
    """从结构化文本中提取根因+对策(改进版)"""
    if not text:
        return {"root_cause": "", "fix_method": "", "raw": ""}

    root_cause = ""
    fix_method = ""

    # 模式1: [root_cause]:xxx / [solution]:xxx
    m = re.search(r'\[root_cause\][：:]\s*(.+?)(?=\n\[|\n\d+#|\Z)', text, re.DOTALL)
    if not m:
        m = re.search(r'\[root_cause\]\s*=\s*(.+?)(?=\n\[|\n\d+#|\Z)', text, re.DOTALL)
    if m:
        root_cause = m.group(1).strip()

    m = re.search(r'\[solution\][：:]\s*(.+?)(?=\n\[|\n\d+#|\Z)', text, re.DOTALL)
    if m:
        fix_method = m.group(1).strip()

    # 模式2: 中文标签
    if not root_cause:
        for label in ("根本原因", "根因分析", "根因", "原因分析", "原因", "问题分析"):
            m = re.search(rf'{label}[：:]\s*(.+?)(?=\n(?:解决|提交|影响|结论|备注|根本|根因)|\n\d+#|\Z)', text, re.DOTALL)
            if m:
                root_cause = m.group(1).strip()
                break

    if not fix_method:
        for label in ("解决对策", "修复方式", "修复方案", "对策", "解决措施", "结论"):
            m = re.search(rf'{label}[：:]\s*(.+?)(?=\n(?:提交|影响|备注|根本|根因)|\n\d+#|\Z)', text, re.DOTALL)
            if m:
                fix_method = m.group(1).strip()
                break

    # 模式3: 提取gerrit/gitlab链接作为修复引用
    if not fix_method:
        links = re.findall(r'(https?://(?:code-gerrit|gitlab)\.desaysv\.com/\S+)', text)
        if links:
            fix_method = f"代码提交: {links[0]}"

    # 模式4: [Patch]:URL
    if not fix_method:
        m = re.search(r'\[Patch\]:\s*(\S+)', text)
        if m:
            fix_method = f"Patch: {m.group(1).strip()}"

    return {"root_cause": root_cause.strip(), "fix_method": fix_method.strip()}


def extract_smart(row: dict) -> dict:
    """智能提取：优先 Cause Analysis → Comments"""
    cause = row.get("Cause Analysis", "")
    comments = row.get("Comments", "")

    # 先尝试 Cause Analysis
    result = extract_structured(cause)

    # 如果没有提取到，把 Cause Analysis 整体作为根因描述
    if not result["root_cause"] and cause:
        # 清理掉纯元数据
        cleaned = re.sub(r'\d+#\s*\w+\(\w+\)创建时间：.*?\n', '', cause)
        cleaned = re.sub(r'修改时间：.*?\n', '', cleaned)
        cleaned = re.sub(r'@\w+\(\w+\)', '', cleaned)
        cleaned = cleaned.strip()
        if cleaned and len(cleaned) > 5:
            result["root_cause"] = cleaned[:300]

    # 如果还没有，尝试 Comments
    if not result["root_cause"]:
        result2 = extract_structured(comments)
        if result2["root_cause"]:
            result["root_cause"] = result2["root_cause"]
        if not result["fix_method"] and result2["fix_method"]:
            result["fix_method"] = result2["fix_method"]

        # 兜底：取 Comments 最后有意义的段落
        if not result["root_cause"] and comments:
            parts = re.split(r'\n\d+[#\']', comments)
            if parts:
                last = parts[-1].strip()
                last = re.sub(r'.*?创建时间：.*?\|.*', '', last)
                last = re.sub(r'.*?修改时间：.*?', '', last)
                last = re.sub(r'@\w+\(\w+\)', '', last)
                lines = [l.strip() for l in last.split('\n') if l.strip() and len(l.strip()) > 10]
                if lines:
                    result["root_cause"] = lines[-1][:300]

    return result


def classify_solution(root_cause: str, fix_method: str, row: dict) -> dict:
    """评估方案合理性"""
    rc = (root_cause + " " + fix_method).lower()
    status = row.get("Status", "").lower()
    comments = row.get("Comments", "").lower()

    rating = "无明确方案"
    issues = []
    positive = []

    # ── 正面信号 ──
    has_gerrit = bool(re.search(r'(gerrit|gitlab)\.desaysv\.com', rc))
    has_mr = bool(re.search(r'mr链接|merge.request|merge_requests', rc))
    has_code_fix = bool(re.search(r'添加|修改|删除|移除|增加|try.catch|判空|反注册|释放|优化', rc))
    has_concrete_step = bool(re.search(r'关闭.*监控|使用.*验证|建议.*排查|排查.*电源|调整.*电流', rc))
    has_merged = bool(re.search(r'重复.*bug|同.*问题|合并.*跟踪|重复.*单|同.*bug', rc))
    has_root_cause = len(root_cause) > 20

    if has_gerrit or has_mr:
        positive.append("有代码提交链接")
    if has_code_fix:
        positive.append("有具体代码修改措施")
    if has_concrete_step:
        positive.append("有可执行的操作步骤")

    # ── 负面信号 ──
    wait_external = bool(re.search(r'等.*(?:释放|修复|版本|更新|回复|分析)|依赖.*(?:修复|解决)|外部问题|三方问题|第三方'
                                    r'|高通.*(?:分析|case|回复)|雄狮.*处理|讯飞.*分析|转.*外部|下个版本|下次解决', rc))
    cannot_analyze = bool(re.search(r'日志不全|无法解析|缺少log|无法复现|没有.*log|无法定位|需要.*复现|需要.*定位|信息不足', rc))
    just_merged = has_merged and not has_code_fix and not has_concrete_step
    only_description = has_root_cause and not has_code_fix and not has_concrete_step and not has_gerrit and not wait_external

    # ── 定级 ──
    if has_gerrit or (has_code_fix and has_root_cause):
        rating = "✅ 合理"
    elif has_concrete_step:
        rating = "✅ 合理(有操作指引)"
    elif has_merged and has_root_cause:
        rating = "➡️ 合并跟踪"
    elif wait_external:
        rating = "⏳ 依赖外部"
        issues.append("需建立外部依赖跟踪机制和SLA")
    elif cannot_analyze:
        rating = "⚠️ 无法定位"
        issues.append("根因未定位，建议增加复现条件和日志抓取")
    elif only_description:
        rating = "⚠️ 仅有根因无对策"
        issues.append("有根因分析但缺少明确的修复方案")
    elif not has_root_cause:
        rating = "❌ 信息缺失"
        issues.append("既无根因也无对策")
    else:
        rating = "⚠️ 仅有根因无对策"

    # Closed 状态 + 有根因，大概率是已解决但方案未记录
    if "closed" in status or "confirm" in status:
        if "无明确方案" in rating or "仅有根因" in rating or "信息缺失" in rating:
            if has_root_cause:
                rating = "⚠️ 方案未记录(已关闭)"
                issues.append("问题已关闭但修复方案未完整记录在系统中")

    return {
        "rating": rating,
        "issues": issues,
        "positive": positive,
    }


# ── 细粒度分类规则 ──
FINE_GRAINED_RULES = [
    {"category": "应用crash-空指针/NPE", "keywords": ["nullpointer", "null pointer", "空指针", "npe",
              "attempt to get length of null", "null object", "null array"]},
    {"category": "应用crash-ANR", "keywords": ["anr", "应用无响应", "not responding",
              "input dispatching timed out", "broadcast of intent"]},
    {"category": "应用crash-Native/SIG", "keywords": ["sigsegv", "sigabrt", "sigbus", "signal 11",
              "signal 6", "signal 7", "native crash", "nativecrash", "fatal exception",
              "fatal signal", "tombstone", "backtrace", "segv_maperr"]},
    {"category": "应用crash-Observer泄漏", "keywords": ["contentobserver", "observer", "未反注册",
              "没有反注册", "register", "unregister", "重复注册", "反复注册", "内存泄漏", "memory leak"]},
    {"category": "应用crash-跨进程/序列化", "keywords": ["跨进程", "序列化", "反序列化", "binder",
              "parcel", "aidl", "ipc"]},
    {"category": "应用crash-三方应用", "keywords": ["com.lion.media", "carota", "讯飞", "雄狮",
              "高德", "百度", "酷我", "三方", "外部问题", "第三方", "第三方应用"]},

    {"category": "QNX-SAIL/safetymonitor", "keywords": ["safetymonitor", "sail", "75ms",
              "md response", "apss", "ramdump", "vsens", "safety_mx", "safety monitor",
              "pshold", "功能安全"]},
    {"category": "QNX-IDPS/Kernel", "keywords": ["idps", "nidps", "kernel crash", "kernel shutdown",
              "qnx crash", "qnx kernel", "kernel panic"]},
    {"category": "QNX-SPI/心跳/通信", "keywords": ["spi", "心跳", "0x80", "0x5501", "握手",
              "heartbeat", "保活", "io-sock", "emac", "驱动挂死", "5501", "通信链路"]},
    {"category": "QNX-启动/STR唤醒", "keywords": ["str", "休眠", "唤醒", "suspend", "resume",
              "sleep", "wakeup", "sr未ready", "sr ready", "sleep mode"]},

    {"category": "硬件-NOC/DDR/900E", "keywords": ["noc error", "ddr", "900e", "0xac12e0",
              "nor error", "高通case", "qualcomm", "硬件问题", "硬件故障", "9008"]},
    {"category": "硬件-显示屏/解串器", "keywords": ["解串器", "serdes", "dsi", "lvds", "显示屏",
              "display", "mipi", "屏幕", "panel", "寄存器", "屏线束", "掉link", "link down"]},
    {"category": "硬件-电源/供电", "keywords": ["电源", "供电", "电流", "逆变器", "电压", "限流",
              "5a", "10a", "power supply", "电源管理", "输入电源", "供电不稳"]},

    {"category": "系统-升级/回滚/配置", "keywords": ["升级", "回滚", "共板", "标定文件", "ota",
              "fota", "刷机", "版本回退", "版本升级", "downgrade", "upgrade", "基线", "升级失败"]},
    {"category": "系统-分区/存储损坏", "keywords": ["userdata", "metadata", "分区", "分区乱码",
              "断电导致", "数据损坏", "文件系统", "fsck", "分区损坏"]},
    {"category": "系统-PAG/动画库", "keywords": ["pag", "动画库", "lottie", "动画", "特效",
              "gpu渲染", "渲染", "opengl", "vulkan", "skia"]},
    {"category": "系统-进程freeze/冻结", "keywords": ["freeze", "冻结", "冻屏", "卡死",
              "system_server", "进程冻结", "process freeze", "程序被freeze"]},
    {"category": "系统-内存踩踏", "keywords": ["踩内存", "memory corruption", "踩踏", "kasan",
              "内存越界", "use after free", "double free", "野指针", "wild pointer"]},
    {"category": "系统-surface/fd泄漏", "keywords": ["surfaceflinger", "surface", "fd泄漏",
              "too many open files", "文件描述符", "buffer", "fence"]},

    {"category": "环境-台架/线束/电源不稳", "keywords": ["台架", "线束", "接触不良", "串口板",
              "adb线", "拔掉线", "电源不稳", "电源波动", "外部电压", "座子浮高"]},
    {"category": "环境-测试手法/工具", "keywords": ["导u盘", "dumpstate", "log导出", "导日志",
              "slog2info", "测试指令", "测试脚本", "monkey", "regulartest", "mtbf"]},
    {"category": "环境-温度/高温", "keywords": ["高温", "温度", "水冷", "散热", "过热", "thermal", "85°"]},

    {"category": "场景-CarPlay/Carlink", "keywords": ["carplay", "carlink", "car life", "hicar",
              "手机互联", "cp连接", "cp回连", "有线cp", "无线cp"]},
    {"category": "场景-倒车/AVM", "keywords": ["倒车", "avm", "r档", "挂r", "倒挡", "rear view",
              "全景", "摄像头", "camera"]},
    {"category": "场景-USB/媒体", "keywords": ["usb", "u盘", "歌曲", "音频", "音乐", "媒体",
              "randomaccessfile", "图库", "图片", "usb加载"]},
]


def classify_issue(row: dict) -> str:
    text = (row.get("Title", "") + " " + row.get("Comments", "") + " " +
            row.get("Cause Analysis", "")).lower()
    best, best_score = "未分类", 0
    for rule in FINE_GRAINED_RULES:
        score = sum(1 for kw in rule["keywords"] if kw.lower() in text)
        if score > best_score:
            best_score = score
            best = rule["category"]
    return best


# ── 主流程 ──
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["sheet1"]
    headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    issues = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            val = ws.cell(r, c).value
            row[h] = str(val).strip() if val else ""
        issues.append(row)

    # 处理每条
    for row in issues:
        extracted = extract_smart(row)
        row["_root_cause"] = extracted["root_cause"]
        row["_fix_method"] = extracted["fix_method"]
        row["_category"] = classify_issue(row)
        ev = classify_solution(extracted["root_cause"], extracted["fix_method"], row)
        row["_rating"] = ev["rating"]
        row["_issues"] = ev["issues"]
        row["_positive"] = ev["positive"]

    # 统计
    cat_counter = Counter(r["_category"] for r in issues)
    rating_counter = Counter(r["_rating"] for r in issues)

    # 分组
    grouped = defaultdict(list)
    for row in issues:
        grouped[row["_category"]].append(row)

    # 生成报告
    lines = []
    lines.append("# 黑卡闪问题深度根因分析报告 (v2 改进版)")
    lines.append(f"\n> 总问题数: **{len(issues)}** | 细粒度分类: {len(FINE_GRAINED_RULES)}类 | 分析时间: 2026-05-09")
    lines.append("")

    # 概览
    lines.append("## 概述")
    has_rc = sum(1 for r in issues if len(r.get("_root_cause", "")) > 10)
    has_fix = sum(1 for r in issues if r.get("_fix_method", ""))
    lines.append(f"- 有根因描述: {has_rc}/{len(issues)} ({has_rc/len(issues)*100:.1f}%)")
    lines.append(f"- 有修复方案: {has_fix}/{len(issues)} ({has_fix/len(issues)*100:.1f}%)")
    lines.append("")

    # 细粒度分类统计
    lines.append("## 一、细粒度根因分类统计")
    lines.append("| 排名 | 根因分类 | 数量 | 占比 |")
    lines.append("|------|----------|------|------|")
    for idx, (cat, cnt) in enumerate(cat_counter.most_common(), 1):
        pct = cnt / len(issues) * 100
        lines.append(f"| {idx} | {cat} | {cnt} | {pct:.1f}% |")
    lines.append("")

    # 对策评估统计
    lines.append("## 二、对策合理性评估")
    lines.append("| 评估结果 | 数量 | 占比 | 说明 |")
    lines.append("|----------|------|------|------|")
    rating_desc = {
        "✅ 合理": "有代码提交或具体修改措施",
        "✅ 合理(有操作指引)": "有可执行的操作步骤",
        "➡️ 合并跟踪": "与已有BUG合并处理",
        "⏳ 依赖外部": "等待第三方/高通/供应商",
        "⚠️ 仅有根因无对策": "有根因分析但缺少修复方案",
        "⚠️ 方案未记录(已关闭)": "已关闭但方案未记录",
        "⚠️ 无法定位": "根因未定位",
        "❌ 信息缺失": "既无根因也无对策",
    }
    for rt, cnt in rating_counter.most_common():
        pct = cnt / len(issues) * 100
        desc = rating_desc.get(rt, "")
        lines.append(f"| {rt} | {cnt} | {pct:.1f}% | {desc} |")
    lines.append("")

    # 交叉表
    cross = defaultdict(lambda: defaultdict(int))
    for row in issues:
        cross[row["_category"]][row["_rating"]] += 1
    lines.append("## 三、分类 × 对策合理性交叉表")
    lines.append("")
    lines.append("| 分类 | 合理 | 合并跟踪 | 依赖外部 | 仅有根因 | 方案未记录 | 无法定位 | 信息缺失 |")
    lines.append("|------|------|----------|----------|----------|------------|----------|----------|")
    for cat in sorted(cross, key=lambda c: sum(cross[c].values()), reverse=True):
        d = cross[cat]
        total = sum(d.values())
        parts = [
            str(d.get("✅ 合理", 0) + d.get("✅ 合理(有操作指引)", 0)),
            str(d.get("➡️ 合并跟踪", 0)),
            str(d.get("⏳ 依赖外部", 0)),
            str(d.get("⚠️ 仅有根因无对策", 0)),
            str(d.get("⚠️ 方案未记录(已关闭)", 0)),
            str(d.get("⚠️ 无法定位", 0)),
            str(d.get("❌ 信息缺失", 0)),
        ]
        lines.append(f"| {cat} | {' | '.join(parts)} |")
    lines.append("")

    # 全量清单
    lines.append("## 四、全量问题清单（按根因分类，含根因+对策+评估）")
    lines.append("")
    for cat in sorted(grouped, key=lambda c: len(grouped[c]), reverse=True):
        bug_list = grouped[cat]
        lines.append(f"### {cat}（{len(bug_list)} 条）")
        lines.append("")

        for idx, bug in enumerate(bug_list, 1):
            bid = bug.get("Bug ID", "")
            title = bug.get("Title", "")[:120]
            status = bug.get("Status", "")
            severity = bug.get("Severity", "")
            module = bug.get("Module", "")
            freq = bug.get("Frequency", "")
            root_cause = bug.get("_root_cause", "")[:300]
            fix_method = bug.get("_fix_method", "")[:300]
            rating = bug.get("_rating", "")
            pos = bug.get("_positive", [])
            iss = bug.get("_issues", [])

            lines.append(f"**{idx}. [{bid}] {title}**")
            lines.append("")
            lines.append(f"> 状态: {status} | 严重度: {severity} | 模块: {module} | 频率: {freq}")
            lines.append(f"> 根因: {root_cause if root_cause else '(未提取到)'}")
            lines.append(f"> 对策: {fix_method if fix_method else '(未提取到)'}")
            lines.append(f"> 评估: **{rating}**")
            if pos:
                lines.append(f"> ✓ {', '.join(pos)}")
            if iss:
                lines.append(f"> ⚠ {', '.join(iss)}")
            lines.append("")

        lines.append("---")
        lines.append("")

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"Done: {OUTPUT_PATH}")
    print(f"Total: {len(issues)} issues, {len(''.join(lines))} chars")
    # Print summary
    print("\n=== Quick Stats ===")
    for rt, cnt in rating_counter.most_common():
        print(f"  {rt}: {cnt} ({cnt/len(issues)*100:.1f}%)")

if __name__ == "__main__":
    main()
