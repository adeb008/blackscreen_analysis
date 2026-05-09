#!/usr/bin/env python
"""黑卡闪问题深度根因分析 —— 全量问题分类 + 对策评估"""

import openpyxl
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Optional

# ── 配置 ──
EXCEL_PATH = r"D:\my_crew\black_screen_data\Bug_20260508171433.xlsx"
OUTPUT_PATH = r"D:\my_crew\black_screen_data\deep_analysis_report.md"

# ── 根因+对策提取 ──
def extract_from_comments(comments: str) -> dict:
    """从 Comments 自由文本中解析根因和修复方式"""
    root_cause = ""
    fix_method = ""

    if not comments:
        return {"root_cause": "", "fix_method": ""}

    # [root_cause]:xxx
    m = re.search(r'\[root_cause\][：:]\s*(.+?)(?=\n\d+#|\Z)', comments, re.DOTALL)
    if not m:
        m = re.search(r'\[root_cause\]\s*=\s*(.+?)(?=\n\d+#|\Z)', comments, re.DOTALL)
    if m:
        root_cause = m.group(1).strip()

    # [solution]:xxx
    m = re.search(r'\[solution\][：:]\s*(.+?)(?=\n\d+#|\Z)', comments, re.DOTALL)
    if m:
        fix_method = m.group(1).strip()

    # [Patch]:URL
    if not fix_method:
        m = re.search(r'\[Patch\]:\s*(\S+)', comments)
        if m:
            fix_method = f"Patch: {m.group(1).strip()}"

    # 中文标签
    if not root_cause:
        for label in ("根本原因", "根因分析", "根因", "原因分析"):
            m = re.search(rf'{label}[：:]\s*(.+?)(?=\n\d+#|\n\d\'#|\Z)', comments, re.DOTALL)
            if m:
                root_cause = m.group(1).strip()
                break

    if not fix_method:
        for label in ("解决对策", "修复方式", "修复方案", "对策", "解决措施"):
            m = re.search(rf'{label}[：:]\s*(.+?)(?=\n\d+#|\n\d\'#|\Z)', comments, re.DOTALL)
            if m:
                fix_method = m.group(1).strip()
                break

    # 问题分析（更低优先级）
    if not root_cause:
        for label in ("问题分析", "问题原因", "原因"):
            m = re.search(rf'{label}[：:]\s*(.+?)(?=\n\d+#|\n\d\'#|\Z)', comments, re.DOTALL)
            if m:
                root_cause = m.group(1).strip()
                break

    # 兜底
    if not root_cause and not fix_method:
        parts = re.split(r'\n\d+[#\']', comments)
        if parts:
            last = parts[-1].strip()
            last = re.sub(r'.*?创建时间：.*?\|.*', '', last)
            last = re.sub(r'.*?修改时间：.*?', '', last)
            last = re.sub(r'@\w+\(\w+\)', '', last)
            lines = [l.strip() for l in last.split('\n') if l.strip() and len(l.strip()) > 10]
            if lines:
                root_cause = lines[-1][:200]

    return {"root_cause": root_cause.strip(), "fix_method": fix_method.strip()}


# ── 细粒度根因分类规则（三级） ──
FINE_GRAINED_RULES = [
    # --- 应用层 crash ---
    {
        "category": "应用层crash-空指针/NPE",
        "keywords": ["nullpointer", "null pointer", "空指针", "npe", "attempt to get length of null",
                      "null object", "null array"],
    },
    {
        "category": "应用层crash-ANR",
        "keywords": ["anr", "应用无响应", "not responding", "input dispatching timed out",
                      "broadcast of intent"],
    },
    {
        "category": "应用层crash-SIG/NativeCrash",
        "keywords": ["sigsegv", "sigabrt", "sigbus", "signal 11", "signal 6", "signal 7",
                      "native crash", "nativecrash", "fatal exception", "fatal signal",
                      "tombstone", "backtrace", "segv_maperr"],
    },
    {
        "category": "应用层crash-Observer泄漏",
        "keywords": ["contentobserver", "observer", "未反注册", "没有反注册", "register",
                      "unregister", "重复注册", "反复注册", "内存泄漏", "memory leak"],
    },
    {
        "category": "应用层crash-跨进程/序列化",
        "keywords": ["跨进程", "序列化", "反序列化", "binder", "parcel", "aidl", "ipc"],
    },
    {
        "category": "应用层crash-三方应用",
        "keywords": ["com.lion.media", "carota", "讯飞", "雄狮", "高德", "百度", "酷我",
                      "三方", "外部问题", "第三方", "music", "第三方应用"],
    },

    # --- QNX/底层 ---
    {
        "category": "QNX-SAIL/safetymonitor",
        "keywords": ["safetymonitor", "sail", "75ms", "md response", "apss", "ramdump",
                      "vsens", "safety_mx", "safety monitor", "pshold", "功能安全"],
    },
    {
        "category": "QNX-IDPS/Kernel",
        "keywords": ["idps", "nidps", "kernel crash", "kernel shutdown", "qnx crash",
                      "qnx kernel", "kernel panic"],
    },
    {
        "category": "QNX-通信/SPI/心跳",
        "keywords": ["spi", "心跳", "0x80", "0x5501", "握手", "heartbeat", "保活",
                      "io-sock", "emac", "驱动挂死", "5501", "通信链路"],
    },
    {
        "category": "QNX-启动/STR",
        "keywords": ["str", "休眠", "唤醒", "suspend", "resume", "sleep", "wakeup",
                      "sr未ready", "sr ready", "sleep mode"],
    },

    # --- 硬件 ---
    {
        "category": "硬件-NOC/DDR/900E",
        "keywords": ["noc error", "ddr", "900e", "0xac12e0", "nor error",
                      "高通case", "qualcomm", "硬件问题", "硬件故障", "ddr问题"],
    },
    {
        "category": "硬件-显示屏/解串器",
        "keywords": ["解串器", "serdes", "dsi", "lvds", "显示屏", "display", "mipi",
                      "屏幕", "panel", "寄存器", "屏线束", "掉link", "link down"],
    },
    {
        "category": "硬件-电源/供电",
        "keywords": ["电源", "供电", "电流", "逆变器", "电压", "限流", "5a", "10a",
                      "power supply", "电源管理", "输入电源", "供电不稳"],
    },

    # --- 软件系统 ---
    {
        "category": "系统-内存问题",
        "keywords": ["踩内存", "memory corruption", "踩踏", "kasan", "内存越界",
                      "use after free", "double free", "wild pointer", "野指针"],
    },
    {
        "category": "系统-进程freeze/冻结",
        "keywords": ["freeze", "冻结", "冻屏", "卡死", "不响应", "系统冻结",
                      "system_server", "进程冻结", "process freeze"],
    },
    {
        "category": "系统-升级/回滚/配置",
        "keywords": ["升级", "回滚", "共板", "标定文件", "ota", "fota", "刷机",
                      "版本回退", "版本升级", "downgrade", "upgrade", "基线"],
    },
    {
        "category": "系统-分区/存储损坏",
        "keywords": ["userdata", "metadata", "分区", "分区乱码", "断电导致",
                      "断电上电", "数据损坏", "文件系统", "fsck", "分区损坏"],
    },
    {
        "category": "系统-第三方库(PAG/动画)",
        "keywords": ["pag", "动画库", "lottie", "动画", "特效", "gpu渲染",
                      "渲染", "opengl", "vulkan", "skia", "surface"],
    },
    {
        "category": "系统-gcore/gdb异常",
        "keywords": ["gcore", "gdb", "dump", "crash dump", "coredump", "tombstone",
                      "抓取dump", "抓取log"],
    },

    # --- 环境/测试 ---
    {
        "category": "环境-台架/线束/电源",
        "keywords": ["台架", "线束", "接触不良", "串口板", "adb线", "拔掉线",
                      "电源不稳", "电源波动", "外部电压"],
    },
    {
        "category": "环境-测试手法/工具干扰",
        "keywords": ["导u盘", "dumpstate", "log导出", "导日志", "slog2info",
                      "测试指令", "测试脚本", "monkey", "regulartest", "mtbf"],
    },
    {
        "category": "环境-温度/高温",
        "keywords": ["高温", "温度", "水冷", "散热", "过热", "thermal",
                      "85°", "温升"],
    },

    # --- 功能场景 ---
    {
        "category": "场景-CarPlay/Carlink",
        "keywords": ["carplay", "carlink", "car life", "hicar", "手机互联",
                      "cp连接", "cp回连", "有线cp", "无线cp"],
    },
    {
        "category": "场景-倒车/AVM",
        "keywords": ["倒车", "avm", "r档", "挂r", "倒挡", "rear view",
                      "全景", "摄像头", "camera"],
    },
    {
        "category": "场景-USB/媒体",
        "keywords": ["usb", "u盘", "歌曲", "音频", "音乐", "媒体", "randomaccessfile",
                      "图库", "图片", "usb加载"],
    },
]


def classify_issue(row: dict) -> dict:
    """细粒度分类"""
    text = (row.get("Title", "") + " " +
            row.get("Comments", "") + " " +
            row.get("Cause Analysis", "")).lower()

    matches = []
    for rule in FINE_GRAINED_RULES:
        score = sum(1 for kw in rule["keywords"] if kw.lower() in text)
        if score > 0:
            matches.append((rule["category"], score))

    matches.sort(key=lambda x: x[1], reverse=True)

    if not matches:
        return {
            "primary": "未分类",
            "all_matches": [],
            "confidence": 0,
        }

    return {
        "primary": matches[0][0],
        "all_matches": [m[0] for m in matches[:3]],
        "confidence": matches[0][1],
    }


def evaluate_solution(row: dict, root_cause: str, fix_method: str) -> dict:
    """评估解决对策的合理性"""
    rc_lower = root_cause.lower()
    fm_lower = fix_method.lower()

    issues = []
    rating = "合理"  # default

    # 检查是否"等第三方/等外部"
    wait_patterns = ["等.*释放", "等.*修复", "等待.*解决", "等.*版本", "等.*更新",
                     "外部问题", "三方问题", "第三方", "转外部", "高通.*分析",
                     "高通case", "高通.*回复", "依赖.*修复", "下次解决", "下个版本"]
    if any(re.search(p, fm_lower) for p in wait_patterns):
        if not any(re.search(p, rc_lower) for p in wait_patterns):
            rating = "待定(依赖外部)"
            issues.append("依赖外部/三方解决，建议建立跟踪机制和SLA")

    # 检查是否只分析了问题没给对策
    no_solution_patterns = ["无", "没有", "暂无", "待定", "待分析", "需要复现", "需要定位",
                            "无法定位", "日志不全", "无法解析", "缺少log", "需要进一步"]
    if not fm_lower or any(p in fm_lower for p in no_solution_patterns):
        if not fm_lower:
            rating = "不合理(无对策)"
            issues.append("没有提供解决对策")
        else:
            rating = "待定(无法定位)"
            issues.append(f"无法定位根因: {fix_method[:80]}")

    # 检查 comment 被截断/不完整
    if comments := row.get("Comments", ""):
        if "创建时间" in comments and len(comments.split("\n")) < 4:
            if not root_cause and not fix_method:
                rating = "信息不足"
                issues.append("Comments 内容不完整，无法获取足够信息")

    # 检查"重复bug"
    if re.search(r"重复.*bug|同.*问题|合并.*跟踪|同.*bug", fm_lower + rc_lower):
        rating = "合理(合并跟踪)"
        issues.append("与已有BUG合并跟踪")

    # 检查解决方案是否有具体链接
    has_gerrit = "gerrit" in fm_lower or "gitlab" in fm_lower or "merge" in fm_lower
    has_patch = "patch" in fm_lower or "commit" in fm_lower or "提交链接" in fm_lower

    if has_gerrit or has_patch:
        if rating == "合理":
            rating = "合理(有代码提交)"

    # 检查是否有具体修改措施
    concrete_actions = ["添加", "修改", "删除", "移除", "增加", "更新", "优化",
                        "try catch", "try-catch", "判空", "反注册", "释放",
                        "关闭", "开启", "调整", "降低", "提高"]
    if any(a in fm_lower for a in concrete_actions):
        if "待定" not in rating:
            rating = "合理(有具体措施)"

    return {
        "rating": rating,
        "issues": issues,
    }


# ── 主流程 ──
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["sheet1"]

    headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    issues = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            val = ws.cell(r, c).value
            row[h] = str(val).strip() if val else ""
        issues.append(row)

    print(f"Total: {len(issues)}")

    # 对每条问题做深度分析
    for row in issues:
        # 提取根因+对策
        comments = row.get("Comments", "")
        cause_analysis = row.get("Cause Analysis", "")

        # 优先从 Cause Analysis 提取
        parsed = extract_from_comments(cause_analysis)
        if not parsed["root_cause"] and not parsed["fix_method"]:
            # 回退到 Comments
            parsed = extract_from_comments(comments)

        row["_parsed_root_cause"] = parsed["root_cause"]
        row["_parsed_fix_method"] = parsed["fix_method"]

        # 细粒度分类
        cls = classify_issue(row)
        row["_fine_category"] = cls["primary"]
        row["_all_categories"] = cls["all_matches"]

        # 评估对策
        eval_result = evaluate_solution(row, parsed["root_cause"], parsed["fix_method"])
        row["_eval_rating"] = eval_result["rating"]
        row["_eval_issues"] = eval_result["issues"]

    # ── 按分类分组 ──
    grouped = defaultdict(list)
    for row in issues:
        cat = row["_fine_category"]
        grouped[cat].append(row)

    # ── 生成 Markdown ──
    lines = []
    lines.append("# 黑卡闪问题深度根因分析报告 (全量)")
    lines.append(f"\n> 分析时间: 2026-05-09 | 总问题数: {len(issues)} | 细粒度分类: {len(FINE_GRAINED_RULES)}类")
    lines.append("")

    # 统计概览
    cat_counter = Counter(r["_fine_category"] for r in issues)
    rating_counter = Counter(r["_eval_rating"] for r in issues)

    lines.append("## 一、细粒度根因分类统计")
    lines.append("")
    lines.append("| 根因分类 | 数量 | 占比 |")
    lines.append("|----------|------|------|")
    for cat, cnt in cat_counter.most_common():
        pct = cnt / len(issues) * 100
        lines.append(f"| {cat} | {cnt} | {pct:.1f}% |")
    lines.append("")

    lines.append("## 二、对策合理性评估总览")
    lines.append("")
    lines.append("| 评估结果 | 数量 | 占比 |")
    lines.append("|----------|------|------|")
    for rt, cnt in rating_counter.most_common():
        pct = cnt / len(issues) * 100
        lines.append(f"| {rt} | {cnt} | {pct:.1f}% |")
    lines.append("")

    # 合理性 x 分类交叉
    cross = defaultdict(lambda: defaultdict(int))
    for row in issues:
        cross[row["_fine_category"]][row["_eval_rating"]] += 1
    lines.append("## 三、分类 × 对策合理性 交叉表")
    lines.append("")
    for cat in sorted(cross, key=lambda c: sum(cross[c].values()), reverse=True):
        parts = [f"{rt}:{cnt}" for rt, cnt in sorted(cross[cat].items(), key=lambda x: x[1], reverse=True)]
        lines.append(f"- **{cat}**: {', '.join(parts)}")
    lines.append("")

    # 全量问题清单（按分类）
    lines.append("## 四、全量问题清单（按根因分类）")
    lines.append("")

    for cat in sorted(grouped, key=lambda c: len(grouped[c]), reverse=True):
        bug_list = grouped[cat]
        lines.append(f"### {cat}（{len(bug_list)} 条）")
        lines.append("")

        for idx, bug in enumerate(bug_list, 1):
            bug_id = bug.get("Bug ID", "")
            title = bug.get("Title", "")[:150]
            status = bug.get("Status", "")
            severity = bug.get("Severity", "")
            module = bug.get("Module", "")
            frequency = bug.get("Frequency", "")
            root_cause = bug.get("_parsed_root_cause", "")[:200]
            fix_method = bug.get("_parsed_fix_method", "")[:200]
            rating = bug.get("_eval_rating", "")
            eval_issues = bug.get("_eval_issues", [])

            lines.append(f"#### {idx}. [{bug_id}] {title}")
            lines.append("")
            lines.append(f"| 属性 | 值 |")
            lines.append(f"|------|----|")
            lines.append(f"| 状态 | {status} |")
            lines.append(f"| 严重度 | {severity} |")
            lines.append(f"| 模块 | {module} |")
            lines.append(f"| 频率 | {frequency} |")
            lines.append(f"| 根因 | {root_cause if root_cause else '未提取到'} |")
            lines.append(f"| 解决对策 | {fix_method if fix_method else '未提取到'} |")
            lines.append(f"| **对策评估** | **{rating}** |")
            if eval_issues:
                for ei in eval_issues:
                    lines.append(f"| ⚠️ 风险 | {ei} |")
            lines.append("")

            if root_cause or fix_method:
                # 合理性判断
                pass

        lines.append("---")
        lines.append("")

    # 写入文件
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"Report written to {OUTPUT_PATH}")
    print(f"Total characters: {len(''.join(lines))}")


if __name__ == "__main__":
    main()
