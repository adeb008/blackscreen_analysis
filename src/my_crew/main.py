#!/usr/bin/env python
"""黑卡闪问题分析 — 双工作流入口

用法:
  crewai run                        # 默认工作流一: 问题分析提炼
  python main.py refine             # 工作流一: 问题分析提炼
  python main.py download           # 工作流二: Analysis 问题下载分析
  python main.py full               # 完整双工作流串联

增量跟踪:
  通过 analyzed_bugs.json 记录已分析的 Bug ID + 状态，
  后续导出只分析新增和状态变更的问题，避免重复。
"""

import json
import sys
import warnings
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from my_crew.crew import MyCrew
from my_crew.tools.bug_knowledge_tool import BugKnowledgeTool, compute_trends
from my_crew.config import get_project_root, get_data_dir, get_kb_path, get_excel_path

warnings.filterwarnings("ignore", category=SyntaxWarning, module="pysbd")

PROJECT_ROOT = get_project_root()
DATA_DIR = get_data_dir()
KB_FILE = get_kb_path()


def _find_latest_excel() -> str:
    """找到 black_screen_data 下最新的 Bug_*.xlsx"""
    path = get_excel_path()
    if path:
        return path
    # 最后的兜底
    return ""


def _incremental_filter(excel_path: str) -> dict:
    """增量过滤: 读取 Excel，过滤掉已分析且状态未变的 Bug"""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]] if rows else []

    # 找 Bug ID 和 Status 列
    bug_id_col = None
    status_col = None
    for i, h in enumerate(headers):
        hl = h.lower().replace(" ", "")
        if "bugid" in hl or "bug id" in hl:
            bug_id_col = i
        if "status" in hl:
            status_col = i

    if bug_id_col is None:
        return {"new": [], "changed": [], "skipped": [], "total": 0,
                "error": "找不到 Bug ID 列"}

    # 构建 issue 列表
    issues = []
    for row in rows[1:]:
        if not row or bug_id_col >= len(row):
            continue
        bid = str(row[bug_id_col]).strip() if row[bug_id_col] else ""
        st = str(row[status_col]).strip() if status_col and status_col < len(row) else ""
        if bid:
            issues.append({"bug_id": bid, "status": st})

    result = BugKnowledgeTool.filter_new_and_changed(issues)

    # 保存本次源文件信息
    kb = BugKnowledgeTool._load_static()
    kb["_meta"]["source_file"] = Path(excel_path).name
    BugKnowledgeTool._save_static(kb)

    result["total"] = len(issues)
    return result


def _filtered_inputs(excel_path: str | None = None) -> tuple[dict, bool]:
    """返回 (inputs_dict, is_already_up_to_date)"""
    import json as _json
    path = excel_path or _find_latest_excel()
    filt = _incremental_filter(path)

    new_count = len(filt.get("new", []))
    changed_count = len(filt.get("changed", []))
    skipped_count = len(filt.get("skipped", []))
    total = filt.get("total", 0)

    print(f"\n{'='*50}")
    print(f"增量过滤: {Path(path).name}")
    print(f"  总计: {total} | 新增: {new_count} | 状态变更: {changed_count} | 跳过: {skipped_count}")
    if filt.get("error"):
        print(f"  ⚠️ {filt['error']}")
    print(f"{'='*50}\n")

    inputs = {
        "topic": "黑卡闪问题提炼分析",
        "current_year": str(datetime.now().year),
        "excel_path": str(Path(path).as_posix()),
        "new_count": str(new_count),
        "changed_count": str(changed_count),
        "skipped_count": str(skipped_count),
        "total_count": str(total),
    }
    return inputs, (new_count == 0 and changed_count == 0)


def run():
    """默认: 工作流一"""
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    inputs, up_to_date = _filtered_inputs(excel_path)
    if up_to_date:
        print("⏭️  无新增/变更 Bug，无需重新分析")
        return
    try:
        MyCrew().refinement_crew().kickoff(inputs=inputs)
    except Exception as e:
        raise Exception(f"工作流一执行失败: {e}")


def refine():
    """工作流一: 问题分析提炼"""
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    inputs, up_to_date = _filtered_inputs(excel_path)
    if up_to_date:
        print("⏭️  无新增/变更 Bug，无需重新分析")
        return
    try:
        MyCrew().refinement_crew().kickoff(inputs=inputs)
    except Exception as e:
        raise Exception(f"工作流一执行失败: {e}")


def refine_complete():
    """工作流一 + LLM精校(需人工判断) + 关键词自学习 + 完整Bug清单附录"""
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    inputs, up_to_date = _filtered_inputs(excel_path)
    
    import subprocess
    base = Path(__file__).parent.parent
    
    if up_to_date:
        print("⏭️  无新增/变更 Bug，跳过 CrewAI，直接执行后处理")
    else:
        try:
            MyCrew().refinement_crew().kickoff(inputs=inputs)
        except Exception as e:
            raise Exception(f"工作流一执行失败: {e}")
    
    # Step 1: LLM 批量精校（只要还有"需人工判断"就精校）
    print("\n>>> LLM 批量精校: 处理需人工判断的 Bug")
    result = subprocess.run(
        ["uv", "run", "python", "scripts/llm_reclassify_manual.py", "--batch", "10"],
        cwd=base, capture_output=True, text=True, timeout=600,
    )
    print(result.stdout)
    if result.returncode != 0:
        print(f"[LLM精校警告] 部分批次失败: {result.stderr[:200]}")
    
    # Step 2: 关键词自学习
    print("\n>>> 关键词自学习: 从 LLM 精校结果提取关键词补充规则")
    result = subprocess.run(
        ["uv", "run", "python", "-c", """
from scripts.learn_keywords_from_llm import learn_keywords
learn_keywords()
"""],
        cwd=base, capture_output=True, text=True, timeout=60,
    )
    print(result.stdout)
    if result.returncode != 0:
        print(f"[关键词学习警告] {result.stderr[:200]}")
    
    # Step 3: 完整 Bug 清单附录
    print("\n>>> 后处理: 生成完整 Bug 清单附录")
    result = subprocess.run(
        ["uv", "run", "python", "scripts/generate_bug_list_appendix.py"],
        cwd=base, capture_output=True, text=True, timeout=120,
    )
    print(result.stdout)
    if result.returncode != 0:
        print(f"[后处理错误] {result.stderr}")
    
    # Step 4: 趋势报告
    print("\n>>> 更新趋势报告")
    result = subprocess.run(
        ["uv", "run", "python", "scripts/trend_heatmap_report.py"],
        cwd=base, capture_output=True, text=True, timeout=30,
    )
    print(result.stdout.strip())
    
    # Step 5: 同步到 Obsidian
    print("\n>>> 同步到 Obsidian 保险库")
    result = subprocess.run(
        ["uv", "run", "python", "scripts/sync_to_obsidian.py"],
        cwd=base, capture_output=True, text=True, timeout=60,
    )
    print(result.stdout)
    if result.returncode != 0:
        print(f"[Obsidian同步警告] {result.stderr[:200]}")


def trends():
    """生成趋势与热力图报告"""
    import json as _json
    from pathlib import Path as _Path
    
    kb_path = KB_FILE
    if not kb_path.exists():
        print("[TREND] 知识库不存在，请先运行工作流一")
        return
    
    # 读取 KB，注入当前轮次快照（如果没有历史记录）
    kb = _json.loads(kb_path.read_text(encoding="utf-8"))
    from collections import Counter
    bugs = kb.get("bugs", {})
    snapshot = {
        "timestamp": __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total": len(bugs),
        "categories": dict(sorted(Counter(
            b.get("category", "未分类") for b in bugs.values()
        ).items(), key=lambda x: -x[1])),
        "modules": dict(sorted(Counter(
            b.get("module", "未知") for b in bugs.values()
        ).items(), key=lambda x: -x[1])[:20]),
        "fix_status": dict(Counter(
            b.get("fix_status", "未知") for b in bugs.values()
        ).most_common()),
        "severity": dict(Counter(
            b.get("severity", "未知") for b in bugs.values()
        ).most_common()),
    }
    history = kb["_meta"].setdefault("run_history", [])
    # 避免重复（同一天同一轮不重复追加）
    if not history or history[-1].get("timestamp", "")[:10] != snapshot["timestamp"][:10]:
        history.append(snapshot)
        if len(history) > 50:
            kb["_meta"]["run_history"] = history[-50:]
        kb_path.write_text(_json.dumps(kb, ensure_ascii=False, indent=2), encoding="utf-8")
    
    print(">>> 生成趋势与热力图报告")
    import subprocess
    result = subprocess.run(
        ["uv", "run", "python", "scripts/trend_heatmap_report.py", "--open"],
        cwd=_Path(__file__).resolve().parent.parent.parent,
        capture_output=True, text=True, timeout=30,
    )
    print(result.stdout)
    if result.returncode != 0:
        print(f"[TREND] 错误: {result.stderr}")


def download():
    """工作流二: Analysis 问题下载分析"""
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    inputs, up_to_date = _filtered_inputs(excel_path)
    if up_to_date:
        print("⏭️  无新增/变更 Bug，无需重新下载分析")
        return
    try:
        MyCrew().download_analysis_crew().kickoff(inputs=inputs)
    except Exception as e:
        raise Exception(f"工作流二执行失败: {e}")


def full():
    """完整双工作流串联"""
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    inputs, up_to_date = _filtered_inputs(excel_path)
    if up_to_date:
        print("⏭️  无新增/变更 Bug，无需重新分析")
        return
    try:
        print(">>> 工作流一: 问题分析提炼")
        MyCrew().refinement_crew().kickoff(inputs=inputs)
        print("\n>>> 工作流二: Analysis 问题下载分析")
        MyCrew().download_analysis_crew().kickoff(inputs=inputs)
    except Exception as e:
        raise Exception(f"完整工作流执行失败: {e}")


def train():
    try:
        inputs, _ = _filtered_inputs()
        MyCrew().crew().train(
            n_iterations=int(sys.argv[1]),
            filename=sys.argv[2],
            inputs=inputs,
        )
    except Exception as e:
        raise Exception(f"训练失败: {e}")


def replay():
    try:
        MyCrew().crew().replay(task_id=sys.argv[1])
    except Exception as e:
        raise Exception(f"回放失败: {e}")


def test():
    try:
        inputs, _ = _filtered_inputs()
        MyCrew().crew().test(
            n_iterations=int(sys.argv[1]),
            eval_llm=sys.argv[2],
            inputs=inputs,
        )
    except Exception as e:
        raise Exception(f"测试失败: {e}")


def run_with_trigger():
    import json
    if len(sys.argv) < 2:
        raise Exception("需要 JSON payload")
    try:
        trigger_payload = json.loads(sys.argv[1])
    except json.JSONDecodeError:
        raise Exception("无效的 JSON payload")
    inputs, _ = _filtered_inputs()
    inputs["crewai_trigger_payload"] = trigger_payload
    try:
        return MyCrew().crew().kickoff(inputs=inputs)
    except Exception as e:
        raise Exception(f"触发执行失败: {e}")
