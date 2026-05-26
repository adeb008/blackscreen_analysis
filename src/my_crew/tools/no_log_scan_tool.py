"""CrewAI Tool — 无日志/无法分析问题扫描工具

从 classification_data.json 或直接从 Excel 扫描以下情况：
1. Actual Result 列为空/无/N/A
2. Actual Result 含无日志相关关键词
3. Solved Scheme/Comments 含无法分析关键词
4. Status=Postpone 且无 Solved Scheme（无法复现+无日志组合）

输出：无日志问题清单 + 分类统计 + 影响评估 + 写入经验库建议
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from crewai.tools import BaseTool
from pydantic import BaseModel, Field


# ============================================================
# 无日志识别关键词
# ============================================================

# Actual Result 字段为空或无效的判断
EMPTY_PATTERNS = [
    "", "无", "n/a", "na", "none", "null", "-", "暂无", "无日志",
    "未填", "待填", "/"
]

# Actual Result 含这些词 → 无日志
ACTUAL_RESULT_NO_LOG_KEYWORDS = [
    "无日志", "no log", "日志缺失", "未抓取到日志", "没有日志",
    "未复现无法抓取", "日志丢失", "未能获取日志",
    "log not found", "无法获取日志", "日志不全",
    "日志未上传", "log丢失", "没抓到日志",
    "日志已清理", "日志过期", "no available log",
    "问题时间点日志", "无问题时间点日志", "问题时间点无日志",
    "日志被冲掉", "日志被冲", "log被冲", "日志冲掉",
]

# Solved Scheme / Comments 含这些词 → 无法分析
CANNOT_ANALYZE_KEYWORDS = [
    "无法分析", "日志不全", "无法定位", "日志缺失", "待抓取日志",
    "等待日志", "无法复现", "未复现", "can't reproduce", "cannot reproduce",
    "无法抓取", "缺少日志", "no log available", "日志不足",
    "无有效日志", "等待复现", "无法复现无日志",
    "当前日志无法分析", "当前log无法分析", "日志无法分析原因",
    "需要再加log分析定位", "需要加log分析", "需要补充log", "需要增加log",
    "需要再加log", "加log定位", "增加log定位", "补充log定位",
    "日志被冲掉", "日志被冲", "log被冲",
    # --- 需要抓ramdump / 保留现场 / 再复现 ---
    "ramdump", "抓dump", "grab dump", "抓ramdump", "dump抓取",
    "保留现场", "现场保留", "保持现场", "保存现场",
    "再复现", "下次复现", "复现后抓", "复现时抓", "等待再次复现",
    "复现抓log", "复现抓取", "复现再抓",
    # --- 加日志/打印 ---
    "加日志", "增加日志", "补充日志", "加打印", "增加打印",
    "加log", "增加log", "加入log", "添加log", "加logging",
    "开启日志", "打开日志", "enable log", "enable logging",
    "需要加日志", "需要增加日志", "需要开日志",
]
# 无日志原因分类
NO_LOG_REASON_RULES = [
    {
        "reason": "未复现/无法复现",
        "keywords": ["未复现", "无法复现", "can't reproduce", "cannot reproduce",
                     "偶现", "低概率", "未复现无法抓取"],
    },
    {
        "reason": "日志已过期/已清理",
        "keywords": ["日志过期", "日志已清理", "log丢失", "日志丢失", "已删除", "过期"],
    },
    {
        "reason": "日志从未上传",
        "keywords": ["未上传", "日志未上传", "没有上传", "未提交日志"],
    },
    {
        "reason": "Actual Result 未填写",
        "keywords": [],   # 兜底：actual_result 为空时
    },
    {
        "reason": "日志被冲掉/覆盖",
        "keywords": ["日志被冲掉", "日志被冲", "log被冲", "日志冲掉",
                     "问题时间点日志", "无问题时间点日志", "问题时间点无日志"],
    },
    {
        "reason": "需要补充Log定位",
        "keywords": ["需要再加log分析定位", "需要加log分析", "需要补充log",
                     "需要增加log", "需要再加log", "加log定位", "增加log定位",
                     "补充log定位", "当前日志无法分析", "当前log无法分析",
                     "日志无法分析原因",
                     "加日志", "增加日志", "补充日志", "加打印", "增加打印",
                     "加log", "增加log", "加入log", "添加log", "加logging",
                     "开启日志", "打开日志", "enable log", "enable logging",
                     "需要加日志", "需要增加日志", "需要开日志"],
    },
    {
        "reason": "需要抓ramdump/保留现场",
        "keywords": ["ramdump", "抓dump", "grab dump", "抓ramdump", "dump抓取",
                     "保留现场", "现场保留", "保持现场", "保存现场",
                     "再复现", "下次复现", "复现后抓", "复现时抓", "等待再次复现",
                     "复现抓log", "复现抓取", "复现再抓"],
    },
]


class NoLogScanInput(BaseModel):
    json_path: str = Field(
        default="outputs/classification_data.json",
        description="classification_data.json 路径（ExcelIssueTool 输出）"
    )
    excel_path: str | None = Field(
        None,
        description="原始 Excel 路径（可选，用于补充 actual_result/solved_scheme 字段）"
    )
    output_path: str = Field(
        default="outputs/no_log_report.md",
        description="输出报告路径"
    )


class NoLogScanTool(BaseTool):
    """扫描 Bug 清单中无日志/无法分析问题，输出专项统计报告"""

    name: str = "no_log_scanner"
    description: str = (
        "Scan classification_data.json (or Excel) to identify bugs with missing/unavailable logs. "
        "Outputs: total count, percentage, reason breakdown, module distribution, "
        "severity distribution, full bug list, and impact assessment. "
        "Input: json_path (classification_data.json), optional excel_path, output_path."
    )
    args_schema: type[BaseModel] = NoLogScanInput

    def _run(self, json_path: str = "outputs/classification_data.json",
             excel_path: str | None = None,
             output_path: str = "outputs/no_log_report.md") -> str:

        # ---- 读取分类数据 ----
        bugs = self._load_bugs(json_path, excel_path)
        if not bugs:
            return f"Error: 未能从 {json_path} 读取任何 Bug 数据"

        total = len(bugs)

        # ---- 识别无日志问题 ----
        no_log_bugs = []
        for bug in bugs:
            result = self._is_no_log(bug)
            if result["is_no_log"]:
                bug["_no_log_reason"] = result["reason"]
                bug["_no_log_evidence"] = result["evidence"]
                no_log_bugs.append(bug)

        no_log_count = len(no_log_bugs)
        no_log_ratio = no_log_count / total * 100 if total > 0 else 0

        # ---- 组装报告 ----
        lines: list[str] = []
        lines.append("# 无日志/无法分析问题专项审计报告")
        lines.append(f"\n> 数据来源: {json_path}")
        lines.append(f"> 总 Bug 数: {total} | 无日志问题数: {no_log_count} | **无日志占比: {no_log_ratio:.1f}%**")
        lines.append(f"> 有效日志覆盖率: **{100 - no_log_ratio:.1f}%**")

        # ---- 1. 原因分布 ----
        lines.append("\n## 1. 无日志原因分布")
        reason_counter: Counter = Counter(b["_no_log_reason"] for b in no_log_bugs)
        lines.append("| 原因 | 数量 | 占无日志比 |")
        lines.append("|------|------|-----------|")
        for reason, cnt in reason_counter.most_common():
            pct = cnt / no_log_count * 100 if no_log_count > 0 else 0
            lines.append(f"| {reason} | {cnt} | {pct:.1f}% |")

        # ---- 2. Module 分布 ----
        lines.append("\n## 2. 按模块分布（无日志问题）")
        mod_counter: Counter = Counter(
            b.get("module") or "未知" for b in no_log_bugs
        )
        lines.append("| 模块 | 无日志数 | 该模块总数 | 无日志率 |")
        lines.append("|------|---------|-----------|---------|")
        mod_total: Counter = Counter(b.get("module") or "未知" for b in bugs)
        for mod, cnt in mod_counter.most_common(15):
            total_mod = mod_total[mod]
            rate = cnt / total_mod * 100 if total_mod > 0 else 0
            lines.append(f"| {mod} | {cnt} | {total_mod} | {rate:.1f}% |")

        # ---- 3. 严重度分布 ----
        lines.append("\n## 3. 按严重度分布（无日志问题）")
        sev_counter: Counter = Counter(
            b.get("severity") or "未知" for b in no_log_bugs
        )
        lines.append("| 严重度 | 无日志数 | 占无日志比 |")
        lines.append("|--------|---------|-----------|")
        for sev, cnt in sev_counter.most_common():
            pct = cnt / no_log_count * 100 if no_log_count > 0 else 0
            lines.append(f"| {sev} | {cnt} | {pct:.1f}% |")

        # ---- 4. 影响评估 ----
        lines.append("\n## 4. 影响评估")
        postpone_bugs = [b for b in no_log_bugs
                         if "postpone" in (b.get("status") or "").lower()
                         or "挂起" in (b.get("status") or "")]
        analysis_bugs = [b for b in no_log_bugs
                         if "analysis" in (b.get("status") or "").lower()]
        open_bugs = [b for b in no_log_bugs
                     if any(w in (b.get("status") or "").lower()
                            for w in ["open", "new", "待处理"])]

        lines.append(f"- **已永久搁置（Postpone）**: {len(postpone_bugs)} 条 — 这些问题大概率不会再跟进日志")
        lines.append(f"- **分析中（Analysis）**: {len(analysis_bugs)} 条 — **需要优先补抓日志**，是工作流二输入的前提")
        lines.append(f"- **待处理（Open/New）**: {len(open_bugs)} 条 — 建议在下次复现时配置持续日志抓取")
        lines.append(f"- **其他状态**: {no_log_count - len(postpone_bugs) - len(analysis_bugs) - len(open_bugs)} 条")

        if analysis_bugs:
            lines.append("\n### 需优先补抓日志的 Analysis 状态问题")
            lines.append("| Bug ID | 问题现象 | 模块 | 无日志原因 |")
            lines.append("|--------|---------|------|-----------|")
            for b in analysis_bugs:
                lines.append(
                    f"| {b.get('bug_id','?')} | {(b.get('title') or '')[:60]} "
                    f"| {b.get('module') or '未知'} | {b.get('_no_log_reason','?')} |"
                )

        # ---- 5. 完整无日志 Bug 清单 ----
        lines.append("\n## 5. 完整无日志问题清单")
        lines.append("| Bug ID | 问题现象 | 模块 | 严重度 | 状态 | 根因分类 | 无日志原因 |")
        lines.append("|--------|---------|------|--------|------|---------|-----------|")
        for b in no_log_bugs:
            lines.append(
                f"| {b.get('bug_id','?')} "
                f"| {(b.get('title') or '')[:50]} "
                f"| {b.get('module') or '未知'} "
                f"| {b.get('severity') or '?'} "
                f"| {b.get('status') or '?'} "
                f"| {b.get('root_cause_category') or '?'} "
                f"| {b.get('_no_log_reason','?')} |"
            )

        # ---- 6. 经验库写入建议 ----
        lines.append("\n## 6. 经验沉淀建议（供经验库更新）")
        self._append_experience_suggestions(lines, no_log_bugs, mod_counter, total)

        # ---- 写文件 ----
        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        report_text = "\n".join(lines)
        out_path.write_text(report_text, encoding="utf-8")

        # ---- 返回摘要给 Agent ----
        summary_lines = [
            f"无日志扫描完成: 总 {total} 条 Bug，无日志 {no_log_count} 条（占比 {no_log_ratio:.1f}%），"
            f"有效日志覆盖率 {100 - no_log_ratio:.1f}%",
            "",
            "原因分布:",
        ]
        for reason, cnt in reason_counter.most_common():
            pct = cnt / no_log_count * 100 if no_log_count > 0 else 0
            summary_lines.append(f"  - {reason}: {cnt} 条 ({pct:.1f}%)")
        summary_lines.append("")
        summary_lines.append(f"需优先补抓日志(Analysis状态): {len(analysis_bugs)} 条")
        summary_lines.append(f"已永久搁置(Postpone): {len(postpone_bugs)} 条")
        summary_lines.append(f"完整报告已写入: {output_path}")

        return "\n".join(summary_lines)

    # ============================================================
    # 识别是否为无日志问题
    # ============================================================

    def _is_no_log(self, bug: dict) -> dict:
        actual = (bug.get("actual_result") or "").strip().lower()
        comments = (bug.get("comments") or "").strip().lower()
        parsed_rc = (bug.get("parsed_root_cause") or "").strip().lower()
        parsed_fix = (bug.get("parsed_fix_method") or "").strip().lower()
        status = (bug.get("status") or "").strip().lower()

        combined_text = f"{actual} {comments} {parsed_rc} {parsed_fix}"

        # 规则1: actual_result 为空或无效
        if actual in EMPTY_PATTERNS or not actual:
            return {
                "is_no_log": True,
                "reason": "Actual Result 未填写",
                "evidence": "actual_result 字段为空/无效",
            }

        # 规则2: actual_result 含无日志关键词
        for kw in ACTUAL_RESULT_NO_LOG_KEYWORDS:
            if kw.lower() in actual:
                return {
                    "is_no_log": True,
                    "reason": self._classify_reason(actual, status),
                    "evidence": f"actual_result 含关键词: {kw}",
                }

        # 规则3: solved_scheme/comments 含无法分析关键词
        for kw in CANNOT_ANALYZE_KEYWORDS:
            if kw.lower() in combined_text:
                return {
                    "is_no_log": True,
                    "reason": self._classify_reason(combined_text, status),
                    "evidence": f"comments/solved_scheme 含关键词: {kw}",
                }

        # 规则4: status=Postpone 且 parsed_root_cause 为空（无法复现+无日志组合）
        if "postpone" in status and not parsed_rc:
            return {
                "is_no_log": True,
                "reason": "未复现/无法复现",
                "evidence": "status=Postpone 且无根因分析内容",
            }

        return {"is_no_log": False, "reason": "", "evidence": ""}

    def _classify_reason(self, text: str, status: str) -> str:
        text_lower = text.lower()
        for rule in NO_LOG_REASON_RULES[:-1]:  # 最后一条是兜底，跳过
            for kw in rule["keywords"]:
                if kw.lower() in text_lower:
                    return rule["reason"]
        if "postpone" in status or "无法复现" in text_lower or "未复现" in text_lower:
            return "未复现/无法复现"
        return "其他/原因不明"

    # ============================================================
    # 加载数据
    # ============================================================

    def _load_bugs(self, json_path: str, excel_path: str | None) -> list[dict]:
        bugs: list[dict] = []

        # 优先从 classification_data.json 读取
        jp = Path(json_path)
        if jp.exists():
            try:
                data = json.loads(jp.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    bugs = data
            except Exception as e:
                print(f"[no_log_scanner] 读取 JSON 失败: {e}")

        # 如果 JSON 里缺 actual_result 字段，尝试从 Excel 补充
        if bugs and excel_path:
            bugs = self._enrich_from_excel(bugs, excel_path)

        return bugs

    def _enrich_from_excel(self, bugs: list[dict], excel_path: str) -> list[dict]:
        """从 Excel 补充 actual_result / solved_scheme 字段"""
        try:
            from openpyxl import load_workbook

            path = Path(excel_path)
            if not path.exists():
                return bugs

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return bugs

            headers = [str(c or "").strip().lower() for c in rows[0]]

            # 找 actual_result / solved_scheme 列
            ar_col = next((i for i, h in enumerate(headers)
                           if "actual result" in h or "actual_result" in h
                           or "实际结果" in h), None)
            ss_col = next((i for i, h in enumerate(headers)
                           if "solved scheme" in h or "solved_scheme" in h
                           or "解决对策" in h or "对策" in h), None)
            id_col = next((i for i, h in enumerate(headers)
                           if h in ["bug id", "bugid", "id", "缺陷id", "问题id"]), None)

            if id_col is None:
                return bugs

            # 建索引
            excel_map: dict[str, dict] = {}
            for row in rows[1:]:
                bid = str(row[id_col] or "").strip()
                if bid:
                    excel_map[bid] = {
                        "actual_result": str(row[ar_col] or "").strip() if ar_col is not None else "",
                        "solved_scheme": str(row[ss_col] or "").strip() if ss_col is not None else "",
                    }

            for bug in bugs:
                bid = bug.get("bug_id", "")
                if bid in excel_map:
                    bug.setdefault("actual_result", excel_map[bid]["actual_result"])
                    bug.setdefault("solved_scheme", excel_map[bid]["solved_scheme"])
        except Exception as e:
            print(f"[no_log_scanner] Excel 补充失败: {e}")

        return bugs

    # ============================================================
    # 经验建议
    # ============================================================

    def _append_experience_suggestions(self, lines: list[str],
                                        no_log_bugs: list[dict],
                                        mod_counter: Counter,
                                        total: int) -> None:
        no_log_count = len(no_log_bugs)
        if no_log_count == 0:
            lines.append("> 本次扫描无日志问题数量为 0，无需写入经验库。")
            return

        lines.append("建议将以下规律写入经验库（由 issue_refiner 调用 ExperienceUpdateTool 执行）：")

        # 高频无日志模块
        top_mods = mod_counter.most_common(3)
        for mod, cnt in top_mods:
            mod_total_in_bugs = sum(1 for b in no_log_bugs if (b.get("module") or "未知") == mod)
            if mod_total_in_bugs >= 2:
                lines.append(
                    f"\n### 建议经验条目: {mod} 模块日志获取率低"
                )
                lines.append(f"- **现象**: {mod} 模块有 {cnt} 条问题无有效日志，占该模块无日志问题前列")
                lines.append(f"- **根因**: 该模块问题偶现率高/NAS路径维护不及时/日志抓取未覆盖该模块")
                lines.append(f"- **建议**: 在 {mod} 相关测试场景中预先配置持续日志抓取，确保每次问题均有完整日志上传")
                lines.append(f"- **关键词**: {mod}, 无日志, 日志缺失")

        # 总体无日志率过高预警
        no_log_ratio = no_log_count / total * 100 if total > 0 else 0
        if no_log_ratio > 20:
            lines.append(f"\n### ⚠️ 预警: 无日志率 {no_log_ratio:.1f}% 超过 20% 警戒线")
            lines.append("- **现象**: 本批次 Bug 中超过 1/5 缺少有效日志")
            lines.append("- **建议**: 检查 NAS 日志上传流程、复现操作规范、以及测试阶段日志抓取覆盖情况")
