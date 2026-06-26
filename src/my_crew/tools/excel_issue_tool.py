"""CrewAI Tool — 黑卡闪问题 Excel 结构化分析（方案A+B 升级版）

细粒度分类 + 权重体系 + 负向关键词排除 + Solved Scheme 列 + LLM二次精校接口"""

from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from crewai.tools import BaseTool
from openpyxl import load_workbook
from pydantic import BaseModel, Field


class ExcelIssueToolInput(BaseModel):
    excel_path: str = Field(..., description="Excel file path, e.g. black_screen_data/Bug_20260514.xlsx (relative to project root)")
    sheet_name: str | None = Field(None, description="Worksheet name. If empty, the first sheet is used.")
    max_examples: int = Field(20, description="Maximum number of typical issue examples to include.")
    tracking_file: str | None = Field(None, description="Path to analyzed_bugs.json for automatic knowledge base persistence. Default: auto-detected from project.")
    json_output: str | None = Field(None, description="Path to write structured classification JSON for downstream tasks. Default: outputs/classification_data.json")


from my_crew.models import FINE_GRAINED_RULES, FALLBACK_CATEGORY, FALLBACK_SECTION

# ============================================================
# 状态判断关键词
# ============================================================
FIXED_STATUS_WORDS = ["closed", "done", "fixed", "verified", "已关闭", "已修复", "已验证", "关闭"]
PENDING_STATUS_WORDS = ["postpone", "open", "new", "挂起", "待处理", "待分析", "依赖"]
UNREPRODUCIBLE_WORDS = ["无法复现", "未复现", "can't reproduce", "cannot reproduce", "日志不全"]
ANALYSIS_STATUS_WORDS = ["analysis"]       # 分析中，无结论
CONFIRM_STATUS_WORDS = ["confirm"]          # 待确认，无结论


class ExcelIssueTool(BaseTool):
    """方案A升级版：细粒度分类 + 权重体系 + 负向关键词排除 + Solved Scheme列纳入分析"""

    name: str = "excel_issue_analyzer"
    description: str = (
        "Read an issue-list Excel file and return a structured markdown summary "
        "including field mapping, distributions, root cause classification (weighted "
        "scoring + negative keyword exclusion), fixed/pending grouping, "
        "cross tables, keyword analysis, and typical examples. "
        "Also includes 'Solved Scheme' and 'Actual Result' columns for comprehensive analysis."
    )
    args_schema: type[BaseModel] = ExcelIssueToolInput

    # 标准字段 → Excel 列名别名（新增 Solved Scheme / Actual Result）
    column_aliases: dict[str, list[str]] = {
        "bug_id": ["bug id", "bugid", "id", "缺陷id", "问题id"],
        "title": ["title", "标题", "问题标题", "summary", "问题现象"],
        "comments": ["comments", "comment", "备注", "评论", "说明", "分析说明"],
        "root_cause": ["root cause", "根因", "根因分析", "原因", "rootcause",
                       "cause analysis", "cause_analysis", "Cause Analysis"],
        "fix_method": ["fix", "fix method", "修复方式", "修复方案", "修复"],
        "solved_scheme": ["solved scheme", "solved_scheme", "解决对策", "对策", "修复对策", "方案"],
        "actual_result": ["actual result", "actual_result", "实际结果", "Actual Result"],
        "status": ["status", "状态", "问题状态", "处理状态", "关闭状态"],
        "severity": ["severity", "严重度", "等级", "优先级", "风险等级"],
        "assignee": ["assignee", "责任人", "负责人", "处理人", "owner"],
        "creator": ["creator", "创建人", "提出人", "提交人", "reporter"],
        "module": ["module", "模块", "系统", "功能", "域", "component"],
        "frequency": ["frequency", "频次", "复现频率", "发生频率", "概率"],
        "analyzer": ["analyzer", "分析人", "分析责任人"],
        "sample_stage": ["sample stage", "样件阶段", "阶段", "sample", "stage"],
    }

    # ========================================================
    # 主入口
    # ========================================================

    def _get_worksheet(self, workbook, sheet_name: str | None):
        """获取工作表，不区分大小写"""
        if sheet_name:
            # 精确匹配优先，再大小写不敏感匹配
            if sheet_name in workbook.sheetnames:
                return workbook[sheet_name]
            lower = sheet_name.lower()
            for name in workbook.sheetnames:
                if name.lower() == lower:
                    return workbook[name]
            # 模糊匹配：sheet_name 出现在任何已有 sheet 名中
            for name in workbook.sheetnames:
                if lower in name.lower() or name.lower() in lower:
                    return workbook[name]
            # 都不匹配，报错
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用: {available}")
        # 默认取第一个 sheet
        return workbook[workbook.sheetnames[0]]

    def _run(self, excel_path: str, sheet_name: str | None = None,
             max_examples: int = 20, tracking_file: str | None = None,
             json_output: str | None = None) -> str:
        path = Path(excel_path)
        if not path.exists():
            return f"Error: Excel file not found: {excel_path}"

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = self._get_worksheet(workbook, sheet_name)
        except ValueError as e:
            return f"Error: {e}"
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return "Error: Excel sheet is empty."

        headers = [self._clean_cell(value) or f"Unnamed Column {index + 1}"
                   for index, value in enumerate(rows[0])]
        records = self._records_from_rows(headers, rows[1:])
        if not records:
            return "Error: Excel sheet has headers but no valid data rows."

        mapping = self._map_columns(headers, records)

        # ---- 对每条记录做加权根因分类（方案A） ----
        for record in records:
            text = self._build_analysis_text(record, mapping)
            evidence_bundle = self._extract_evidence_bundle(record, mapping)
            rule_candidates = self._generate_rule_candidates(evidence_bundle)
            rc = self._classify_weighted(text)

            top_candidate = (rule_candidates.get("top_candidates") or [{}])[0]
            final_category = rule_candidates.get("final_candidate") or FALLBACK_CATEGORY
            final_section = self._section_for_category(final_category)
            final_score = top_candidate.get("score", 0)
            final_keywords = top_candidate.get("matched_keywords", [])

            if final_category == FALLBACK_CATEGORY and rc["category"] != FALLBACK_CATEGORY:
                final_category = rc["category"]
                final_section = rc["section"]
                final_score = rc["score"]
                final_keywords = rc["matched_keywords"]

            record["_evidence_bundle"] = evidence_bundle
            record["_rule_candidates"] = rule_candidates
            record["_root_cause_category"] = final_category
            record["_root_cause_section"] = final_section or rc["section"]
            record["_root_cause_score"] = str(final_score)
            record["_root_cause_matched"] = ", ".join(final_keywords)

            # 从 Comments / Cause Analysis / Solved Scheme 提取根因和修复方式
            dedicated_rc = self._value(record, mapping.get("root_cause"))
            dedicated_solved = self._value(record, mapping.get("solved_scheme"))
            parsed = self._parse_comments(self._value(record, mapping.get("comments")))
            record["_parsed_root_cause"] = dedicated_rc or parsed["root_cause"]
            record["_parsed_fix_method"] = dedicated_solved or parsed["fix_method"]

            # 判断修复状态
            record["_fix_status"] = self._classify_fix_status(record, mapping, text)

        # ---- 增量状态标记 + 知识库自动回写（闭环） ----
        if tracking_file:
            self._annotate_incremental_status(records, mapping, tracking_file)
            self._persist_to_kb(records, mapping, tracking_file, str(path))
        else:
            for record in records:
                record.setdefault("_incremental_status", "full")

        # ---- 结构化 JSON 输出（供下游 Agent 精确引用） ----
        if json_output:
            self._write_classification_json(records, mapping, json_output)

        # ---- 组装输出 ----
        lines: list[str] = []
        lines.append("# 黑卡闪问题 Excel 结构化分析摘要（方案A+B）")
        lines.append(f"- 文件: {excel_path}")
        lines.append(f"- Sheet: {worksheet.title}")
        lines.append(f"- 数据行数: {len(records)}")
        lines.append("- 分类体系: 细粒度权重评分 + 负向关键词排除")

        # ---- 字段映射 ----
        lines.append("\n## 字段映射")
        for std_name in ["bug_id", "title", "comments", "root_cause", "fix_method",
                         "solved_scheme", "actual_result",
                         "status", "severity", "assignee", "creator", "module",
                         "frequency", "analyzer", "sample_stage"]:
            match = mapping.get(std_name)
            lines.append(f"- {std_name}: {match or '未识别'}")

        # ---- 字段缺失 ----
        lines.append("\n## 字段缺失情况")
        for column, count in self._missing_counts(headers, records).most_common(10):
            ratio = count / len(records) * 100
            lines.append(f"- {column}: {count} ({ratio:.1f}%)")

        # ---- 标准分布 ----
        self._append_distribution(lines, "Status 分布", records, mapping.get("status"))
        self._append_distribution(lines, "Severity 分布", records, mapping.get("severity"))
        self._append_distribution(lines, "Module 分布", records, mapping.get("module"))
        self._append_distribution(lines, "Frequency 分布", records, mapping.get("frequency"))

        # ---- 根因分类统计 ----
        lines.append("\n## 根因分类统计（细粒度）")
        rc_counter: Counter = Counter()
        rc_section_map: dict[str, set[str]] = defaultdict(set)
        for record in records:
            cat = record.get("_root_cause_category", FALLBACK_CATEGORY)
            rc_counter[cat] += 1
            section = record.get("_root_cause_section", FALLBACK_SECTION)
            rc_section_map[cat].add(section)

        for cat, count in rc_counter.most_common():
            sections = ", ".join(rc_section_map[cat])
            ratio = count / len(records) * 100
            lines.append(f"- {cat}: {count} ({ratio:.1f}%) [{sections}]")

        # ---- 修复状态分组 ----
        lines.append("\n## 修复状态分组")
        fix_counter = Counter(r.get("_fix_status", "未知") for r in records)
        for status, count in fix_counter.most_common():
            ratio = count / len(records) * 100
            lines.append(f"- {status}: {count} ({ratio:.1f}%)")

        # ---- 各根因大类下的 Bug 清单 ----
        lines.append("\n## 各根因大类 Bug 清单")
        bugs_by_cat: dict[str, list[dict]] = defaultdict(list)
        for record in records:
            bugs_by_cat[record.get("_root_cause_category", FALLBACK_CATEGORY)].append(record)

        for cat in sorted(bugs_by_cat, key=lambda c: len(bugs_by_cat[c]), reverse=True):
            bug_list = bugs_by_cat[cat]
            lines.append(f"\n### {cat}（{len(bug_list)} 条）")
            for bug in bug_list[:15]:  # 每类展示 15 条供 LLM 参考，全量数据从 JSON 读取
                bug_id = self._value(bug, mapping.get("bug_id")) or "无 ID"
                title = self._value(bug, mapping.get("title")) or "无标题"
                rc = (bug.get("_parsed_root_cause")
                      or self._value(bug, mapping.get("root_cause"))
                      or "无根因")
                fix = (bug.get("_parsed_fix_method")
                       or self._value(bug, mapping.get("fix_method"))
                       or self._value(bug, mapping.get("solved_scheme"))
                       or "无修复方式")
                status = self._value(bug, mapping.get("status")) or "无状态"
                keywords = bug.get("_root_cause_matched", "")
                lines.append(f"  - [{bug_id}] {title}")
                lines.append(f"    - 根因: {rc[:200]}")
                lines.append(f"    - 修复: {fix[:200]}")
                lines.append(f"    - 状态: {status} | 匹配关键词: {keywords[:100]}")

        # ---- 交叉统计 ----
        self._append_cross_table(lines, "Module x Severity", records,
                                 mapping.get("module"), mapping.get("severity"))
        self._append_cross_table(lines, "Module x 根因分类", records,
                                 mapping.get("module"), "_root_cause_category")
        self._append_cross_table(lines, "根因分类 x 修复状态", records,
                                 "_root_cause_category", "_fix_status")

        # ---- 典型样例 ----
        lines.append(f"\n## 典型问题样例（前 {max_examples} 条）")
        for index, record in enumerate(records[:max_examples], 1):
            bug_id = self._value(record, mapping.get("bug_id")) or "无 Bug ID"
            title = self._value(record, mapping.get("title")) or "无 Title"
            severity = self._value(record, mapping.get("severity")) or "无 Severity"
            status = self._value(record, mapping.get("status")) or "无 Status"
            rc_cat = record.get("_root_cause_category", FALLBACK_CATEGORY)
            fix_st = record.get("_fix_status", "未知")
            comments = self._value(record, mapping.get("comments"))
            root_cause = (record.get("_parsed_root_cause")
                          or self._value(record, mapping.get("root_cause")))
            fix_method = (record.get("_parsed_fix_method")
                          or self._value(record, mapping.get("fix_method"))
                          or self._value(record, mapping.get("solved_scheme")))
            lines.append(
                f"{index}. [{bug_id}] {title} | "
                f"根因分类={rc_cat} | 修复状态={fix_st} | Severity={severity} | Status={status}"
            )
            if root_cause:
                lines.append(f"   - 根因: {root_cause[:200]}")
            if fix_method:
                lines.append(f"   - 修复: {fix_method[:200]}")
            elif comments:
                lines.append(f"   - 备注: {comments[:200]}")

        # ---- 给后续 Agent 的分析提示（含精校指引） ----
        lines.append("\n## 后续 Agent 分析提示")
        lines.append("- 根因分类: 每条 Bug 已按加权规则自动分类，field `_root_cause_category`")
        lines.append(f"- 权重评分: 每条 Bug 的匹配评分记录在 `_root_cause_score`")
        lines.append("- 修复状态: 已分入 已修复 / 未修复/挂起 / 无法复现，field `_fix_status`")
        lines.append("- 匹配关键词: 每条 Bug 命中的关键词记录在 `_root_cause_matched`")
        lines.append("- 方案B精校: issue_refiner 需对每条的 `_root_cause_category` 做 LLM 二次审核，")
        lines.append("  修正因否定语义、文本矛盾造成的误分类，输出结构化精校结果")
        lines.append("- 报告结构建议: 按「已修复→未修复→统计→经验→总结」五段式组织")

        return "\n".join(lines)

    # ========================================================
    # 方案A：加权根因分类
    # ========================================================

    def _build_analysis_text(self, record: dict[str, str],
                             mapping: dict[str, str | None]) -> str:
        """构建分析文本：Title + Comments + Cause Analysis + Solved Scheme
        （不包含 Actual Result — 那是工作流二的输入：问题描述+时间+日志路径）"""
        source_keys = ["title", "comments", "root_cause", "solved_scheme"]
        parts = []
        for key in source_keys:
            col = mapping.get(key)
            if col and record.get(col):
                parts.append(record[col])
        return "\n".join(parts)

    def _classify_weighted(self, text: str) -> dict:
        """加权根因分类（方案A核心算法）

        算法:
          1. 对每条规则，在 text 中匹配各级关键词
          2. 如果任何 exclude_keyword 命中，跳过该规则
          3. 计分: strong=3分/个, medium=2分/个, weak=1分/个
          4. 最高分者胜出；同分时 priority 高者胜出
          5. 全不命中 → FALLBACK_CATEGORY
        """
        text_lower = text.lower()
        best = {
            "category": FALLBACK_CATEGORY,
            "section": FALLBACK_SECTION,
            "score": 0,
            "matched_keywords": [],
            "conflict_hint": "",
        }

        for rule in FINE_GRAINED_RULES:
            # 负向排除检查
            excluded = False
            for ekw in rule.get("exclude_keywords", []):
                if ekw.lower() in text_lower:
                    excluded = True
                    break
            if excluded:
                continue

            # 加权计分
            score = 0
            hits = []
            for strength, weight in [("strong", 3), ("medium", 2), ("weak", 1)]:
                for kw in rule["keywords"].get(strength, []):
                    if kw.lower() in text_lower:
                        score += weight
                        hits.append(kw)

            if score == 0:
                continue

            # 比较：更高分胜出；同分按 priority
            if score > best["score"] or (
                score == best["score"]
                and rule["priority"] > (best.get("_priority", 0))
            ):
                best = {
                    "category": rule["name"],
                    "section": rule["section"],
                    "score": score,
                    "matched_keywords": hits,
                    "_priority": rule["priority"],
                    "conflict_hint": "",
                }
            elif score == best["score"] and score > 0:
                # 同分但未胜出，记录冲突提示
                best["conflict_hint"] = (
                    f"与「{rule['name']}」同分({score}分)，按优先级取当前分类"
                )

        return best

    def _extract_evidence_bundle(self, record: dict[str, str],
                                 mapping: dict[str, str | None]) -> dict[str, list[dict[str, Any]]]:
        """按字段拆分证据，区分症状/根因/修复/上下文/否定证据。"""
        evidence: dict[str, list[dict[str, Any]]] = {
            "symptom": [],
            "root_cause": [],
            "fix": [],
            "context": [],
            "negative": [],
        }

        field_specs = [
            ("title", "symptom", 1.0),
            ("comments", "context", 1.2),
            ("root_cause", "root_cause", 3.2),
            ("fix_method", "fix", 1.0),
            ("solved_scheme", "fix", 0.8),
        ]

        for field_key, default_bucket, weight in field_specs:
            column = mapping.get(field_key)
            text = self._value(record, column)
            if not text:
                continue
            source_field = field_key if field_key != "root_cause" else "cause_analysis"
            fragments = self._split_evidence_fragments(text)
            if not fragments:
                fragments = [text.strip()]
            for fragment in fragments:
                bucket = self._infer_evidence_bucket(fragment, default_bucket)
                evidence[bucket].append({
                    "text": fragment,
                    "source_field": source_field,
                    "weight": weight,
                })

        parsed = self._parse_comments(self._value(record, mapping.get("comments")))
        if parsed.get("root_cause"):
            evidence["root_cause"].append({
                "text": parsed["root_cause"],
                "source_field": "comments",
                "weight": 2.0,
            })
        if parsed.get("fix_method"):
            evidence["fix"].append({
                "text": parsed["fix_method"],
                "source_field": "comments",
                "weight": 1.5,
            })

        return {
            key: self._deduplicate_evidence_items(items)
            for key, items in evidence.items()
        }

    def _generate_rule_candidates(self,
                                  evidence_bundle: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
        """基于字段级证据重新计算候选分类，支持否定证据抵消。"""
        top_candidates: list[dict[str, Any]] = []
        excluded_candidates: list[dict[str, Any]] = []
        all_positive_text = "\n".join(
            item["text"]
            for bucket in ("symptom", "root_cause", "fix", "context")
            for item in evidence_bundle.get(bucket, [])
        ).lower()

        for rule in FINE_GRAINED_RULES:
            excluded_by = [
                keyword for keyword in rule.get("exclude_keywords", [])
                if keyword.lower() in all_positive_text
            ]
            if excluded_by:
                excluded_candidates.append({
                    "category": rule["name"],
                    "excluded_by": excluded_by,
                })
                continue

            field_hits: dict[str, list[str]] = defaultdict(list)
            negative_hits: dict[str, list[str]] = defaultdict(list)
            matched_keywords: list[str] = []
            score = 0.0

            for bucket, sign in (("symptom", 1), ("root_cause", 1), ("fix", 1), ("context", 1), ("negative", -1)):
                for item in evidence_bundle.get(bucket, []):
                    text_lower = item["text"].lower()
                    source_field = item["source_field"]
                    field_weight = float(item.get("weight", 1.0))
                    for strength, keyword_weight in (("strong", 3), ("medium", 2), ("weak", 1)):
                        for keyword in rule["keywords"].get(strength, []):
                            if keyword.lower() not in text_lower:
                                continue
                            delta = field_weight * keyword_weight * sign
                            score += delta
                            if sign > 0:
                                field_hits[source_field].append(keyword)
                                matched_keywords.append(keyword)
                            else:
                                negative_hits[source_field].append(keyword)

            score = round(score, 2)
            if score <= 0:
                continue

            top_candidates.append({
                "category": rule["name"],
                "score": score,
                "field_hits": {key: sorted(set(value)) for key, value in field_hits.items()},
                "negative_hits": {key: sorted(set(value)) for key, value in negative_hits.items()},
                "matched_keywords": sorted(set(matched_keywords)),
                "excluded_by": [],
                "priority": rule["priority"],
            })

        top_candidates.sort(key=lambda item: (item["score"], item.get("priority", 0)), reverse=True)
        final_candidate = top_candidates[0]["category"] if top_candidates else FALLBACK_CATEGORY
        confidence = self._estimate_candidate_confidence(top_candidates)
        needs_llm_review = (
            final_candidate == FALLBACK_CATEGORY
            or confidence < 0.6
            or (len(top_candidates) >= 2 and (top_candidates[0]["score"] - top_candidates[1]["score"] <= 1.5))
        )

        return {
            "final_candidate": final_candidate,
            "confidence": confidence,
            "top_candidates": top_candidates[:5],
            "excluded_candidates": excluded_candidates[:5],
            "needs_llm_review": needs_llm_review,
        }

    def _estimate_candidate_confidence(self, top_candidates: list[dict[str, Any]]) -> float:
        if not top_candidates:
            return 0.0
        best = float(top_candidates[0]["score"])
        second = float(top_candidates[1]["score"]) if len(top_candidates) > 1 else 0.0
        if best <= 0:
            return 0.0
        confidence = best / (best + second + 1.0)
        return round(min(0.99, confidence), 2)

    def _section_for_category(self, category: str) -> str:
        for rule in FINE_GRAINED_RULES:
            if rule["name"] == category:
                return rule.get("section", FALLBACK_SECTION)
        return FALLBACK_SECTION

    def _split_evidence_fragments(self, text: str) -> list[str]:
        cleaned = self._strip_noise_lines(text)
        fragments = re.split(r"[\n；;。]+", cleaned)
        return [fragment.strip() for fragment in fragments if fragment and fragment.strip()]

    def _strip_noise_lines(self, text: str) -> str:
        if not text:
            return ""
        cleaned_lines: list[str] = []
        for raw_line in str(text).splitlines():
            line = raw_line.strip()
            if not line:
                continue
            lower = line.lower()
            if (
                re.search(r"uid[a-z0-9]+", lower)
                or "创建时间" in line
                or "修改时间" in line
                or line.startswith("log路径")
                or lower.startswith("thanks")
                or line.startswith("\\\\")
                or "desaysv.com" in lower
                or "hzhhnnas01" in lower
            ):
                continue
            cleaned_lines.append(line)
        return "\n".join(cleaned_lines)

    def _infer_evidence_bucket(self, text: str, default_bucket: str) -> str:
        text_lower = text.lower()
        negative_patterns = ["没有发现", "未发现", "无异常", "排除", "不是", "非", "not ", "no "]
        root_cause_patterns = ["根因", "原因", "导致", "引起", "系", "定位", "分析", "确认"]
        fix_patterns = ["修复", "解决", "优化", "修改", "增加", "patch", "方案", "对策"]
        symptom_patterns = ["黑屏", "花屏", "闪屏", "重启", "卡死", "无声", "异常", "问题"]

        if any(pattern in text_lower for pattern in negative_patterns):
            return "negative"
        if any(pattern in text_lower for pattern in fix_patterns):
            return "fix"
        if any(pattern in text_lower for pattern in root_cause_patterns):
            return "root_cause"
        if any(keyword in text_lower for keyword in ("高通", "qualcomm", "vendor", "供应商")):
            return "root_cause"
        if any(pattern in text_lower for pattern in symptom_patterns):
            return "symptom"
        return default_bucket

    def _deduplicate_evidence_items(self, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        merged: dict[tuple[str, str], dict[str, Any]] = {}
        for item in items:
            key = (item["source_field"], item["text"])
            current = merged.get(key)
            if current is None or item.get("weight", 0) > current.get("weight", 0):
                merged[key] = item
        return list(merged.values())

    def _classify_fix_status(self, record: dict, mapping: dict, text: str) -> str:
        """判断修复状态（新增 Analysis/Confirm 特殊处理）"""
        status_val = self._value(record, mapping.get("status")).lower()
        text_lower = text.lower()
        if any(w in text_lower for w in UNREPRODUCIBLE_WORDS):
            return "无法复现"
        if any(w in status_val for w in PENDING_STATUS_WORDS):
            return "未修复/挂起"
        if any(w in status_val for w in FIXED_STATUS_WORDS):
            return "已修复"
        # Analysis / Confirm 没有分析结论，不视为已修复
        if any(w in status_val for w in ANALYSIS_STATUS_WORDS):
            return "未修复（分析中）"
        if any(w in status_val for w in CONFIRM_STATUS_WORDS):
            return "未修复（待确认）"
        # fallback: 按 section 推断
        section = record.get("_root_cause_section", "")
        if "未修复" in section or "挂起" in section:
            return "未修复/挂起"
        return "已修复"

    # ========================================================
    # 知识库持久化（闭环回写）
    # ========================================================

    def _persist_to_kb(self, records: list[dict], mapping: dict,
                       tracking_file: str, source_excel: str) -> None:
        """将每条 Bug 的分类结果写入 analyzed_bugs.json"""
        import json as _json
        from datetime import datetime as _datetime

        kb_path = Path(tracking_file)
        kb_path.parent.mkdir(parents=True, exist_ok=True)

        # 加载现有知识库
        kb: dict = {"_meta": {}, "bugs": {}}
        if kb_path.exists():
            try:
                kb = _json.loads(kb_path.read_text(encoding="utf-8"))
            except (_json.JSONDecodeError, OSError):
                pass

        now = _datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        written = 0
        for record in records:
            bug_id = self._value(record, mapping.get("bug_id"))
            if not bug_id:
                continue

            entry = kb.get("bugs", {}).get(bug_id, {})
            entry["category"] = record.get("_root_cause_category", "")
            entry["fix_status"] = record.get("_fix_status", "")
            entry["refined_at"] = now
            # 保留已有的工作流二字段
            if self._value(record, mapping.get("status")):
                entry["status"] = self._value(record, mapping.get("status"))
            if self._value(record, mapping.get("severity")):
                entry["severity"] = self._value(record, mapping.get("severity"))
            if self._value(record, mapping.get("module")):
                entry["module"] = self._value(record, mapping.get("module"))
            if self._value(record, mapping.get("title")):
                entry["title"] = self._value(record, mapping.get("title"))
            kb["bugs"][bug_id] = entry
            written += 1

        # 更新元信息
        now = _datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        kb["_meta"]["last_run"] = now
        kb["_meta"]["source_file"] = Path(source_excel).name
        kb["_meta"]["total_analyzed"] = len(kb.get("bugs", {}))

        # 趋势统计 + 历史快照
        cat_fixed: dict[str, int] = {}
        cat_total: dict[str, int] = {}
        mod_ct: dict[str, int] = {}
        for b in kb.get("bugs", {}).values():
            cat = b.get("category", "未分类")
            mod = b.get("module", "未知")
            cat_total[cat] = cat_total.get(cat, 0) + 1
            mod_ct[mod] = mod_ct.get(mod, 0) + 1
            if b.get("fix_status") == "已修复" or b.get("status", "").lower() in ("closed", "confirm"):
                cat_fixed[cat] = cat_fixed.get(cat, 0) + 1

        category_trend = {}
        for cat, total in sorted(cat_total.items(), key=lambda x: -x[1]):
            fixed = cat_fixed.get(cat, 0)
            rate = fixed / total * 100 if total > 0 else 0
            if rate >= 80:
                trend = "✅ 收敛"
            elif rate >= 50:
                trend = "🔶 收敛中"
            else:
                trend = "🔴 需关注"
            category_trend[cat] = {"total": total, "fixed": fixed,
                                    "fix_rate": round(rate, 1), "trend": trend}
        kb["_meta"]["category_trend"] = category_trend
        kb["_meta"]["module_heatmap"] = dict(
            sorted(mod_ct.items(), key=lambda x: -x[1])[:15])

        # 历史快照（供趋势+热力图使用）
        snapshot = {
            "timestamp": now,
            "total": len(kb.get("bugs", {})),
            "categories": dict(sorted(cat_total.items(), key=lambda x: -x[1])),
            "modules": dict(sorted(mod_ct.items(), key=lambda x: -x[1])[:20]),
            "fix_status": {},
            "severity": {},
        }
        history = kb["_meta"].setdefault("run_history", [])
        history.append(snapshot)
        if len(history) > 50:
            kb["_meta"]["run_history"] = history[-50:]

        kb_path.write_text(_json.dumps(kb, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[知识库] 已回写 {written} 条 Bug 分类记录到 {tracking_file}")

    def _annotate_incremental_status(self, records: list[dict], mapping: dict,
                                     tracking_file: str) -> None:
        """在写回知识库前，根据既有 KB 标记 new / changed / skipped。"""
        import json as _json
        force_full = os.getenv("FORCE_FULL_RUN") == "1"

        if force_full:
            for record in records:
                record["_incremental_status"] = "new"
            return

        kb_path = Path(tracking_file)
        existing_bugs: dict[str, dict[str, Any]] = {}
        if kb_path.exists():
            try:
                existing_bugs = (_json.loads(kb_path.read_text(encoding="utf-8")) or {}).get("bugs", {})
            except (_json.JSONDecodeError, OSError, AttributeError):
                existing_bugs = {}

        for record in records:
            bug_id = self._value(record, mapping.get("bug_id"))
            current_status = self._value(record, mapping.get("status")).strip().lower()
            if not bug_id:
                record["_incremental_status"] = "new"
                continue

            existing = existing_bugs.get(bug_id)
            if not existing:
                record["_incremental_status"] = "new"
                continue

            old_status = str(existing.get("status", "")).strip().lower()
            record["_incremental_status"] = "changed" if old_status != current_status else "skipped"

    def _write_classification_json(self, records: list[dict], mapping: dict,
                                    json_path: str) -> None:
        """输出结构化分类数据 JSON，供下游 Agent 精确引用"""
        import json as _json

        items = []
        for record in records:
            bug_id = self._value(record, mapping.get("bug_id"))
            if not bug_id:
                continue
            project_value = os.environ.get("PROJECT_SCHEMA", "").strip() or self._value(record, mapping.get("project"))
            incremental_status = record.get("_incremental_status", "")
            items.append({
                "bug_id": bug_id,
                "project": project_value,
                "incremental_status": incremental_status,
                "title": self._value(record, mapping.get("title")),
                "status": self._value(record, mapping.get("status")),
                "severity": self._value(record, mapping.get("severity")),
                "module": self._value(record, mapping.get("module")),
                "raw_fields": {
                    "title": self._value(record, mapping.get("title")),
                    "comments": self._value(record, mapping.get("comments")),
                    "root_cause": self._value(record, mapping.get("root_cause")),
                    "fix_method": self._value(record, mapping.get("fix_method")),
                    "solved_scheme": self._value(record, mapping.get("solved_scheme")),
                    "status": self._value(record, mapping.get("status")),
                    "severity": self._value(record, mapping.get("severity")),
                    "module": self._value(record, mapping.get("module")),
                },
                "evidence": record.get("_evidence_bundle", {
                    "symptom": [],
                    "root_cause": [],
                    "fix": [],
                    "context": [],
                    "negative": [],
                }),
                "rule_engine": record.get("_rule_candidates", {
                    "final_candidate": record.get("_root_cause_category", FALLBACK_CATEGORY),
                    "confidence": 0.0,
                    "top_candidates": [],
                    "excluded_candidates": [],
                    "needs_llm_review": True,
                }),
                "final": {
                    "root_cause_category": record.get("_root_cause_category", ""),
                    "root_cause_section": record.get("_root_cause_section", ""),
                    "fix_status": record.get("_fix_status", ""),
                    "score": record.get("_root_cause_score", "0"),
                    "matched_keywords": record.get("_root_cause_matched", ""),
                    "parsed_root_cause": record.get("_parsed_root_cause", ""),
                    "parsed_fix_method": record.get("_parsed_fix_method", ""),
                    "llm_confidence": record.get("_llm_confidence", ""),
                    "decision_reason": record.get("_decision_reason", ""),
                },
            })

        payload = {
            "meta": {
                "version": "2.0",
                "item_count": len(items),
                "schema": "evidence-rule-engine-v2",
            },
            "items": items,
        }
        output = _json.dumps(payload, ensure_ascii=False, indent=2)
        Path(json_path).parent.mkdir(parents=True, exist_ok=True)
        Path(json_path).write_text(output, encoding="utf-8")
        print(f"[JSON] 已输出 {len(items)} 条结构化分类数据到 {json_path}")

    # ========================================================
    # Comments 解析（保持原逻辑，增强 Solved Scheme 支持）
    # ========================================================

    def _parse_comments(self, comments: str) -> dict:
        """从 Comments / Cause Analysis / Solved Scheme 提取根因和修复方式。

        支持的格式:
          1. [root_cause]:xxx / [solution]:xxx / [Patch]:URL
          2. 根本原因:xxx / 解决对策:xxx / 问题分析:xxx / 原因:xxx
          3. 最后一条有意义的整句（兜底）
        """
        if not comments:
            return {"root_cause": "", "fix_method": ""}

        comments = self._strip_noise_lines(comments)

        root_cause = ""
        fix_method = ""

        # 方括号标签格式
        m = re.search(r'\[root_cause\][：:]\s*(.+?)(?:\n\d+#|\Z)', comments, re.DOTALL)
        if not m:
            m = re.search(r'\[root_cause\]\s*=\s*(.+?)(?:\n\d+#|\Z)', comments, re.DOTALL)
        if m:
            root_cause = m.group(1).strip()

        m = re.search(r'\[solution\][：:]\s*(.+?)(?:\n\d+#|\Z)', comments, re.DOTALL)
        if m:
            fix_method = m.group(1).strip()

        if not fix_method:
            m = re.search(r'\[Patch\]:\s*(\S+)', comments)
            if m:
                fix_method = f"Patch: {m.group(1).strip()}"

        # 中文标签（方括号未命中时兜底）
        if not root_cause:
            for label in ("根本原因", "根因分析", "根因", "原因分析"):
                m = re.search(
                    rf'{label}[：:]\s*(.+?)(?:\n(?:解决|提交|影响|结论|备注|根本|根因)|\n\d+#|\Z)',
                    comments, re.DOTALL)
                if m:
                    root_cause = m.group(1).strip()
                    break

        if not fix_method:
            for label in ("解决对策", "修复方式", "修复方案", "对策"):
                m = re.search(
                    rf'{label}[：:]\s*(.+?)(?:\n(?:提交|影响|备注|根本|根因)|\n\d+#|\Z)',
                    comments, re.DOTALL)
                if m:
                    fix_method = m.group(1).strip()
                    break

        # 更低优先级：问题分析/原因
        if not root_cause:
            for label in ("问题分析", "原因"):
                m = re.search(
                    rf'{label}[：:]\s*(.+?)(?:\n\d+#|\Z)', comments, re.DOTALL)
                if m:
                    root_cause = m.group(1).strip()
                    break

        # 兜底：取最后有意义的段落
        if not root_cause and not fix_method:
            parts = re.split(r'\n\d+[#\']', comments)
            if parts:
                last = parts[-1].strip()
                last = re.sub(r'.*?创建时间：.*?\|.*', '', last)
                last = re.sub(r'.*?修改时间：.*?', '', last)
                last = re.sub(r'@\w+\(\w+\)', '', last)
                lines = [l.strip() for l in last.split('\n')
                         if l.strip() and len(l.strip()) > 10]
                if lines:
                    root_cause = lines[-1][:200]

        return {"root_cause": root_cause.strip(), "fix_method": fix_method.strip()}

    # ========================================================
    # 数据处理助手
    # ========================================================

    def _records_from_rows(self, headers: list[str],
                           rows: list[tuple[Any, ...]]) -> list[dict[str, str]]:
        records: list[dict[str, str]] = []
        for row in rows:
            record = {}
            for index, header in enumerate(headers):
                record[header] = self._clean_cell(
                    row[index] if index < len(row) else None)
            if any(record.values()):
                records.append(record)
        return records

    def _map_columns(self, headers: list[str],
                     records: list[dict[str, str]]) -> dict[str, str | None]:
        mapping = {key: self._guess_column_by_name(headers, aliases)
                   for key, aliases in self.column_aliases.items()}
        if mapping.get("title") is None:
            mapping["title"] = self._guess_long_text_column(headers, records)
        if mapping.get("comments") is None:
            mapping["comments"] = self._guess_long_text_column(
                headers, records, exclude={mapping.get("title")})
        if mapping.get("status") is None:
            mapping["status"] = self._guess_status_column(headers, records)
        if mapping.get("severity") is None:
            mapping["severity"] = self._guess_severity_column(headers, records)
        return mapping

    def _guess_column_by_name(self, headers: list[str],
                              aliases: list[str]) -> str | None:
        normalized_headers = [(header, self._normalize(header)) for header in headers]
        normalized_aliases = [self._normalize(alias) for alias in aliases]
        for header, normalized_header in normalized_headers:
            for alias in normalized_aliases:
                if alias and (
                    alias == normalized_header
                    or alias in normalized_header
                    or normalized_header in alias
                ):
                    return header
        return None

    def _guess_long_text_column(self, headers: list[str],
                                records: list[dict[str, str]],
                                exclude: set[str | None] | None = None) -> str | None:
        exclude = exclude or set()
        best_column = None
        best_score = 0.0
        for header in headers:
            if header in exclude:
                continue
            values = [record.get(header, "") for record in records if record.get(header)]
            if not values:
                continue
            avg_len = sum(len(value) for value in values) / len(values)
            unique_ratio = len(set(values)) / len(values)
            score = avg_len * 0.7 + unique_ratio * 20
            if score > best_score:
                best_score = score
                best_column = header
        return best_column if best_score >= 15 else None

    def _guess_status_column(self, headers: list[str],
                             records: list[dict[str, str]]) -> str | None:
        words = ["open", "closed", "done", "fixed", "reject", "new", "active",
                 "处理中", "已关闭", "关闭", "待处理", "已解决"]
        return self._guess_by_content_words(headers, records, words, threshold=0.25)

    def _guess_severity_column(self, headers: list[str],
                               records: list[dict[str, str]]) -> str | None:
        words = ["critical", "major", "minor", "blocker", "high", "medium", "low",
                 "严重", "一般", "高", "中", "低", "s", "p"]
        return self._guess_by_content_words(headers, records, words, threshold=0.2)

    def _guess_by_content_words(self, headers: list[str],
                                records: list[dict[str, str]],
                                words: list[str], threshold: float) -> str | None:
        for header in headers:
            values = [record.get(header, "").lower() for record in records if record.get(header)]
            if not values:
                continue
            hit_count = sum(any(word in value for word in words) for value in values)
            if hit_count / len(values) >= threshold:
                return header
        return None

    def _missing_counts(self, headers: list[str],
                        records: list[dict[str, str]]) -> Counter:
        return Counter({
            header: sum(1 for record in records if not record.get(header))
            for header in headers
        })

    def _append_distribution(self, lines: list[str], title: str,
                             records: list[dict[str, str]],
                             column: str | None) -> None:
        if not column:
            lines.append(f"\n## {title}\n- 未识别对应字段")
            return
        lines.append(f"\n## {title}")
        for value, count in self._counts(records, column).most_common(15):
            ratio = count / len(records) * 100
            lines.append(f"- {value}: {count} ({ratio:.1f}%)")

    def _append_cross_table(self, lines: list[str], title: str,
                            records: list[dict[str, str]],
                            row_column: str | None,
                            value_column: str | None) -> None:
        if not row_column or not value_column:
            lines.append(f"\n## {title}\n- 未识别足够字段")
            return
        lines.append(f"\n## {title}")
        grouped: dict[str, Counter] = defaultdict(Counter)
        for record in records:
            row_value = record.get(row_column) or "空值"
            value = record.get(value_column) or "空值"
            grouped[row_value][value] += 1
        for row_value, counter in sorted(
            grouped.items(), key=lambda item: sum(item[1].values()), reverse=True
        )[:12]:
            parts = [f"{v}: {c}" for v, c in counter.most_common(8)]
            lines.append(f"- {row_value}: " + "; ".join(parts))

    def _counts(self, records: list[dict[str, str]],
                column: str) -> Counter:
        return Counter(record.get(column) or "空值" for record in records)

    def _value(self, record: dict[str, str], column: str | None) -> str:
        return record.get(column, "") if column else ""

    def _clean_cell(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize(self, value: str) -> str:
        return "".join(str(value).lower()
                       .replace("_", " ")
                       .replace("-", " ")
                       .split())
