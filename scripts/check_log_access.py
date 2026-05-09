#!/usr/bin/env python
"""检查 Analysis 问题的日志路径访问权限"""
import openpyxl, re, subprocess

EXCEL = r"D:\my_crew\black_screen_data\Bug_20260509113654.xlsx"

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["sheet1"]
headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

idx = {h: i for i, h in enumerate(headers)}

issues = []
for r in range(2, ws.max_row + 1):
    status = str(ws.cell(r, idx["Status"] + 1).value or "")
    if "analysis" not in status.lower():
        continue

    ar = str(ws.cell(r, idx["Actual Result"] + 1).value or "")
    log_path = ""
    m = re.search(r"日志链接[：:]\\s*(\\\\\S+)", ar)
    if not m:
        m = re.search(r"日志地址[：:]\\s*(\\\\\S+)", ar)
    if not m:
        m = re.search(r"log链接[：:]\\s*(\\\\\S+)", ar)
    if m:
        raw = m.group(1).strip()
        raw = re.sub(r"[。，,;]+$", "", raw)
        log_path = raw

    issues.append({
        "bug_id": str(ws.cell(r, idx["Bug ID"] + 1).value or ""),
        "title": str(ws.cell(r, idx["Title"] + 1).value or "")[:120],
        "status": status,
        "severity": str(ws.cell(r, idx["Severity"] + 1).value or ""),
        "module": str(ws.cell(r, idx["Module"] + 1).value or ""),
        "freq": str(ws.cell(r, idx["Frequency"] + 1).value or ""),
        "assignee": str(ws.cell(r, idx["Assignee"] + 1).value or ""),
        "log_path": log_path,
    })

total = len(issues)
accessible = []
denied = []
no_path = []

for i, iss in enumerate(issues):
    label = f"[{i+1}/{total}] {iss['bug_id']}"
    
    if not iss["log_path"]:
        print(f"{label} -> NO PATH in Actual Result")
        no_path.append(iss)
        continue

    # Convert UNC path for bash: \\server\share -> //server/share
    p = iss["log_path"]
    # Replace \ with /
    p = p.replace("\\", "/")
    # Ensure // prefix
    if not p.startswith("//"):
        p = "//" + p.lstrip("/")

    try:
        r = subprocess.run(["ls", p], capture_output=True, text=True, timeout=10)
        if r.returncode == 0 and r.stdout.strip():
            count = len(r.stdout.strip().split("\n"))
            print(f"{label} -> OK ({count} items)")
            accessible.append({**iss, "item_count": count, "unc_path": p})
        elif r.returncode == 0:
            print(f"{label} -> EMPTY")
            denied.append({**iss, "reason": "目录为空", "unc_path": p})
        elif "denied" in r.stderr.lower() or "not found" in r.stderr.lower():
            print(f"{label} -> DENIED")
            denied.append({**iss, "reason": r.stderr.strip()[:100], "unc_path": p})
        else:
            print(f"{label} -> ERR: {r.stderr.strip()[:80]}")
            denied.append({**iss, "reason": r.stderr.strip()[:100], "unc_path": p})
    except subprocess.TimeoutExpired:
        print(f"{label} -> TIMEOUT")
        denied.append({**iss, "reason": "访问超时", "unc_path": p})
    except Exception as e:
        print(f"{label} -> EXCEPTION: {e}")
        denied.append({**iss, "reason": str(e)[:100], "unc_path": p})

print()
print("=" * 60)
print(f"SUMMARY: Total={total} | OK={len(accessible)} | DENIED={len(denied)} | NO_PATH={len(no_path)}")
print("=" * 60)

if accessible:
    print("\n=== ACCESSIBLE ===")
    for iss in accessible:
        print(f"  [{iss['bug_id']}] {iss['assignee']} | Severity={iss['severity']} | {iss['item_count']} items")
        print(f"    Title: {iss['title']}")
        print(f"    Path:  {iss['log_path']}")
        print()

if denied:
    print("\n=== DENIED / ERROR (需开权限) ===")
    for iss in denied:
        print(f"  [{iss['bug_id']}] {iss['assignee']} | Severity={iss['severity']}")
        print(f"    Title: {iss['title']}")
        print(f"    Path:  {iss['log_path']}")
        print(f"    Reason: {iss.get('reason', 'unknown')}")
        print()

if no_path:
    print("\n=== NO LOG PATH (Actual Result 缺少日志链接) ===")
    for iss in no_path:
        print(f"  [{iss['bug_id']}] {iss['assignee']} | Severity={iss['severity']}")
        print(f"    Title: {iss['title']}")
        print()
