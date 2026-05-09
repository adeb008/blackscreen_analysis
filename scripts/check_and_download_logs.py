#!/usr/bin/env python
"""检查 Analysis 问题的日志路径访问权限，有权限则下载"""

import openpyxl, re, subprocess, os, shutil, unicodedata

EXCEL = r"D:\my_crew\black_screen_data\Bug_20260509113654.xlsx"
LOG_DIR = r"D:\my_crew\T1Q黑卡闪问题分析"

def sanitize(s):
    """清理文件名中的非法字符"""
    s = s.replace("/", "_").replace("\\", "_").replace(":", "_")
    s = s.replace("*", "_").replace("?", "_").replace("\"", "_")
    s = s.replace("<", "_").replace(">", "_").replace("|", "_")
    return s[:60]

def extract_log_path(actual_result):
    """从 Actual Result 列提取日志 UNC 路径"""
    if not actual_result:
        return ""
    
    ar = str(actual_result)
    
    # 找到"日志链接："之后的内容
    m = re.search(r'日志链接[：:]\\s*(.+)', ar)
    if not m:
        m = re.search(r'日志地址[：:]\\s*(.+)', ar)
    if not m:
        m = re.search(r'log链接[：:]\\s*(.+)', ar)
    
    if not m:
        return ""
    
    raw = m.group(1).strip()
    
    # 截断到下一个中文标签或换行
    raw = re.split(r'[\n\r]|出现问题时间|版本链接|视频链接', raw)[0].strip()
    # 去掉尾部标点
    raw = re.sub(r'[。，,;]+$', '', raw)
    
    # Excel 中UNC路径的\被存储为\\，归一化
    # \\\\hzhhnnas01\\DIDA6003... → \\hzhhnnas01\DIDA6003...
    normalized = raw.replace("\\\\", "\\")
    
    # 确保以 \\ 开头
    if not normalized.startswith("\\\\"):
        if normalized.startswith("\\"):
            normalized = "\\" + normalized
        else:
            normalized = "\\\\" + normalized.lstrip("\\")
    
    return normalized

# ── 读取 Excel ──
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["sheet1"]
headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

idx_bugid = headers.index("Bug ID")
idx_title = headers.index("Title")
idx_status = headers.index("Status")
idx_ar = headers.index("Actual Result")
idx_assignee = headers.index("Assignee")
idx_severity = headers.index("Severity")
idx_module = headers.index("Module")
idx_freq = headers.index("Frequency")

issues = []
for r in range(2, ws.max_row + 1):
    status = str(ws.cell(r, idx_status + 1).value or "")
    if "analysis" not in status.lower():
        continue
    
    ar_val = ws.cell(r, idx_ar + 1).value
    log_path = extract_log_path(ar_val)
    
    issues.append({
        "bug_id": str(ws.cell(r, idx_bugid + 1).value or ""),
        "title": str(ws.cell(r, idx_title + 1).value or ""),
        "severity": str(ws.cell(r, idx_severity + 1).value or ""),
        "module": str(ws.cell(r, idx_module + 1).value or ""),
        "assignee": str(ws.cell(r, idx_assignee + 1).value or ""),
        "log_path": log_path,
    })

# ── 逐个检查并下载 ──
total = len(issues)
print(f"Analysis 问题总数: {total}")
print(f"日志存储目录: {LOG_DIR}")
print()

accessible = []
denied = []
no_path = []

for i, iss in enumerate(issues):
    bid = iss["bug_id"]
    label = f"[{i+1}/{total}] {bid}"
    
    if not iss["log_path"]:
        print(f"{label} -> SKIP: Actual Result 中无日志路径")
        no_path.append(iss)
        continue
    
    # 转换 UNC 路径为 bash 可访问格式
    unc_windows = iss["log_path"]  # \\hzhhnnas01.desaysv.com\DIDA6003\...
    unc_bash = unc_windows.replace("\\", "/")
    # 确保 // 前缀
    if not unc_bash.startswith("//"):
        unc_bash = "//" + unc_bash.lstrip("/")
    while "//" in unc_bash[2:]:
        unc_bash = unc_bash[:2] + unc_bash[2:].replace("//", "/")
    
    # 测试访问
    print(f"{label} -> 测试: {unc_bash[:100]}...")
    
    try:
        r = subprocess.run(["ls", unc_bash], capture_output=True, text=True, timeout=15)
        if r.returncode == 0 and r.stdout.strip():
            items = [x for x in r.stdout.strip().split("\n") if x]
            count = len(items)
            print(f"  -> OK ({count} items), 开始下载...")
            
            # 创建本地目录: bugid_描述_日志
            safe_title = sanitize(iss["title"])
            dirname = f"{bid}_{safe_title}"
            local_dir = os.path.join(LOG_DIR, dirname)
            os.makedirs(local_dir, exist_ok=True)
            
            # 递归复制
            try:
                # Use cp -r for recursive copy
                src = unc_bash.rstrip("/") + "/."
                cp_result = subprocess.run(
                    ["cp", "-r", src, local_dir],
                    capture_output=True, text=True, timeout=300
                )
                if cp_result.returncode == 0:
                    # Count downloaded files
                    file_count = 0
                    for root, dirs, files in os.walk(local_dir):
                        file_count += len(files)
                    print(f"  -> DOWNLOADED: {file_count} files -> {local_dir}")
                    accessible.append({**iss, "local_dir": local_dir, "file_count": file_count,
                                      "unc_path": unc_windows, "remote_items": count})
                else:
                    print(f"  -> COPY FAILED: {cp_result.stderr[:100]}")
                    denied.append({**iss, "reason": f"复制失败: {cp_result.stderr[:100]}",
                                  "unc_path": unc_windows})
            except subprocess.TimeoutExpired:
                print(f"  -> TIMEOUT during copy")
                denied.append({**iss, "reason": "下载超时", "unc_path": unc_windows})
                
        elif r.returncode == 0:
            print(f"  -> EMPTY (目录为空)")
            denied.append({**iss, "reason": "目录为空", "unc_path": unc_windows})
        else:
            err = r.stderr.strip()
            if "denied" in err.lower() or "not found" in err.lower():
                print(f"  -> DENIED: 无访问权限")
                denied.append({**iss, "reason": "无访问权限", "unc_path": unc_windows})
            else:
                print(f"  -> ERROR: {err[:100]}")
                denied.append({**iss, "reason": err[:100], "unc_path": unc_windows})
    except subprocess.TimeoutExpired:
        print(f"  -> TIMEOUT")
        denied.append({**iss, "reason": "连接超时", "unc_path": unc_windows})
    except Exception as e:
        print(f"  -> EXCEPTION: {e}")
        denied.append({**iss, "reason": str(e)[:100], "unc_path": unc_windows})
    
    print()

# ── 汇总 ──
print("=" * 70)
print(f"汇总: 总计={total} | 已下载={len(accessible)} | 无权限={len(denied)} | 无路径={len(no_path)}")
print("=" * 70)

if accessible:
    print(f"\n{'='*40}")
    print("已下载 (可进入下一步分析):")
    print(f"{'='*40}")
    for iss in accessible:
        print(f"  [{iss['bug_id']}] {iss['assignee']} | {iss['severity']}")
        print(f"    标题: {iss['title'][:80]}")
        print(f"    远程: {iss['unc_path']}")
        print(f"    本地: {iss['local_dir']} ({iss['file_count']} files)")
        print()

if denied:
    print(f"\n{'='*40}")
    print("需开权限:")
    print(f"{'='*40}")
    for iss in denied:
        print(f"  [{iss['bug_id']}] {iss['assignee']} | {iss['severity']}")
        print(f"    标题: {iss['title'][:80]}")
        print(f"    路径: {iss.get('unc_path', 'N/A')}")
        print(f"    原因: {iss.get('reason', 'unknown')}")
        print()

if no_path:
    print(f"\n{'='*40}")
    print("Actual Result 中缺少日志链接:")
    print(f"{'='*40}")
    for iss in no_path:
        print(f"  [{iss['bug_id']}] {iss['assignee']} | {iss['severity']}")
        print(f"    标题: {iss['title'][:80]}")
        print()
